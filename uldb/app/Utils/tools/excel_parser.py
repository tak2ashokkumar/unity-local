# -*- coding: utf-8 -*-
"""
    excel_parser

    :copyright: (C) 2016 UnitedLayer, LLC. All Rights Reserved.
    :author: rtapia@unitedlayer.com
"""
from __future__ import absolute_import
from openpyxl import load_workbook

CATEGORY_START = 'A1'
CATEGORY_END = 'A100'


def get_category_offset(ws, name):
    for row in ws.iter_rows('%s:%s' % (CATEGORY_START, CATEGORY_END)):
        if row[0].value == name:
            return row[0].row
    return None


def get_column_offset(ws, row_num, name):
    # print row_num, name
    for col in ws.rows[row_num - 1]:
        if col.value == name:
            return col.col_idx
    return None


def read_row(ws, row_num, fields):
    r = {'name': ws.cell(row=row_num, column=1).value}
    r.update({k: ws.cell(row=row_num, column=v).value for k, v in fields.iteritems()})
    return r


def get_data(ws, name, cols, payload):
    offset = get_category_offset(ws, name)
    fields = {field: get_column_offset(ws, offset, field) for field in cols}
    i = 1
    while ws.cell(column=1, row=(offset + i)).value:
        d = read_row(ws, offset + i, fields)
        payload[name].append(d)
        i += 1
    return payload


_config = {
    'ranges': {
        'CUSTOMER_INFO': 'Customer',
        'VCENTERS': 'vCenters',
        'HYPERVISORS': 'Hypervisors',
        'BARE METALS': 'Bare Metals',
        'SANs': 'SANs',
        'LOAD_BALANCERS': 'Load Balancers',
        'FIREWALLS': 'Firewalls',
        'SWITCHES': 'Switches',
        'CAGES': 'Cages',
        'CABINETS': 'Cabinets',
        'PDUs': 'PDUs',
    }
}


def parse_excel(filename):
    wb = load_workbook(filename)
    result = {}
    for key, sheet_name in _config['ranges'].iteritems():
        sheet = wb.get_sheet_by_name(sheet_name)
        values = [[cell.value for cell in row] for row in sheet.iter_rows()]
        # print values
        if values:
            key_list = [x.replace(" ", "") for x in values[0] if x is not None]
            result[key] = []
            for row in values[1:]:
                if row:
                    if any(row[:len(key_list)]):
                        result[key].append(dict(zip(key_list, row[:len(key_list)])))
    return result
