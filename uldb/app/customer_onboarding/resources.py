import sys
import uuid
from datetime import datetime

import openpyxl

from cloud.hyperv.models import HypervVM

if sys.version_info[0] > 2:
    from io import BytesIO
else:
    from cStringIO import StringIO as BytesIO

from copy import deepcopy
from import_export import resources
from import_export.instance_loaders import ModelInstanceLoader
from import_export import fields
from import_export.widgets import ForeignKeyWidget, ManyToManyWidget, BooleanWidget

from django.contrib.contenttypes.models import ContentType

from app.datacenter.models import *
from app.inventory.models import *
from cloud.CloudService.models import ColoCloud, Cloud, PrivateCloud
from cloud.vmware.models import VMwareVcenter
from cloud.VmwareAdapter.models import VmwareVmMigration
from cloud.vcloud.models import VCloudVirtualMachines
from cloud.proxmox.models import ProxmoxVM
from cloud.openstack_app.models import OpenStackVmMigration
from app.server.models import Instance
from app.organization.models import Organization
from integ.networking.models import *
from .fields import ManufacturerWidgetWithCreation, OperatingSystemForeignKeyWidget, \
    VcenterURLWidget


class ErrorLoggingModelResource(resources.ModelResource):
    delete_field = fields.Field(
        column_name='Delete',
        attribute='delete_field',
    )

    def import_row(self, row, instance_loader, **kwargs):
        from import_export.results import RowResult
        # overriding import_row to ignore errors and skip rows that fail to import
        # without failing the entire import
        import_result = super(ErrorLoggingModelResource, self).import_row(row, instance_loader, **kwargs)
        if import_result.import_type == RowResult.IMPORT_TYPE_ERROR:
            # Copy the values to display in the preview report
            import_result.diff = [row[val] for val in row]
            # Add a column with the error message
            import_result.diff.append('Errors: {}'.format([err.error for err in import_result.errors]))
            # clear errors and mark the record to skip
            import_result.errors = []
            import_result.import_type = RowResult.IMPORT_TYPE_SKIP

        return import_result

    def for_delete(self, row, instance):
        if self.fields['delete_field'].clean(row) == "Delete Asset":
            return True
        return False

    def before_import(self, dataset, using_transactions, dry_run, **kwargs):
        id_field_index = self._meta.id_field_index
        id_feild_col = dataset.get_col(id_field_index)

        customer_id = kwargs['customer_id']
        id_list = []
        uuid_list = []

        id_feild = self._meta.id_field_name
        model_class = self._meta.model
        for each_value in id_feild_col:
            look_ups = {
                id_feild: each_value
            }
            if hasattr(self._meta, 'customer_feild'):
                customer_feild = self._meta.customer_feild
                look_ups[customer_feild] = customer_id
            try:
                model_object = model_class.objects.get(**look_ups)
                id_list.append(model_object.id)
            except model_class.DoesNotExist:
                id_list.append(None)

            if hasattr(model_class, 'uuid'):
                try:
                    model_object = model_class.objects.get(**look_ups)
                    uuid_list.append(model_object.uuid)
                except model_class.DoesNotExist:
                    uuid_list.append(None)

        customer_col = [customer_id] * dataset.height

        if id_list:
            dataset.append_col(id_list, 'id')
        if hasattr(self._meta.model, 'uuid'):
            if uuid_list:
                dataset.append_col(uuid_list, 'uuid')

        if hasattr(self._meta, 'customer_feild'):
            if customer_col:
                dataset.append_col(customer_col, self._meta.customer_feild)

    def after_import(self, dataset, result, using_transactions, dry_run, **kwargs):
        from copy import deepcopy
        file_path = kwargs.get('file_path', None)
        if file_path:
            _file = open(file_path)
            excel_file = openpyxl.reader.excel.load_workbook(BytesIO(_file.read()))
            excel_sheet = excel_file[dataset.title]
            for index, row in enumerate(result.rows):
                errors_out = ''
                if row.errors:
                    for each_error in row.errors:
                        errors_out += each_error.error.args[0].split('\n')[0]
                    
                    excel_sheet.cell(row=index + 3, column=self.status_column).value = "Error"
                    excel_sheet.cell(row=index + 3, column=self.status_column + 1).value = str(errors_out)
                else:
                    excel_sheet.cell(row=index + 3, column=self.status_column).value = "Success"
                    excel_sheet.cell(row=index + 3, column=self.status_column + 1).value = ""

            excel_file.save(kwargs['file_path'])

    @classmethod
    def content_type_id(self):
        content_type_id = ContentType.objects.get_for_model(
            self._meta.model
        ).id
        return content_type_id


class ColoCloudResource(ErrorLoggingModelResource):

    def for_delete(self, row, instance):
        return False

    name = fields.Field(
        column_name='Name',
        attribute='name',
    )

    location = fields.Field(
        column_name='Location',
        attribute='location',
    )

    customer_id = fields.Field(
        column_name='customer_id',
        attribute='customer_id',
    )
    status_column = 3

    class Meta:
        model = ColoCloud
        id_field_index = 0
        id_field_name = 'name'
        customer_feild = 'customer_id'


class CabinetResource(ErrorLoggingModelResource):
    name = fields.Field(
        column_name='Name',
        attribute='name',
    )

    cabinet_type = fields.Field(
        column_name='Type',
        attribute='cabinet_type',
        widget=ForeignKeyWidget(CabinetTypes, 'cabinet_type')
    )

    customers = fields.Field(
        column_name='customers',
        attribute="customers",
        widget=ManyToManyWidget(Organization, 'id')
    )

    colocloud_set = fields.Field(
        column_name='Datacenter',
        attribute="colocloud_set",
        widget=ManyToManyWidget(ColoCloud, field='name')
    )

    cabinet_model = fields.Field(
        column_name='Model',
        attribute='cabinet_model',
        widget=ForeignKeyWidget(CabinetModels, 'model')
    )

    status_column = 6

    class Meta:
        model = Cabinet
        id_field_index = 0
        id_field_name = 'name'
        customer_feild = 'customers'


class PDUResource(ErrorLoggingModelResource):
    name = fields.Field(
        column_name='Name',
        attribute='name',
    )

    cabinet = fields.Field(
        column_name='Cabinet',
        attribute='cabinet',
        widget=ForeignKeyWidget(Cabinet, 'name')
    )

    model = fields.Field(
        column_name='Model',
        attribute='model',
        widget=ForeignKeyWidget(PDUModel, 'model_number')
    )

    pdu_type = fields.Field(
        column_name='PDU Type',
        attribute='pdu_type',
    )

    position = fields.Field(
        column_name='Position',
        attribute='position',
        default=0
    )

    sockets = fields.Field(
        column_name='Number of sockets',
        attribute='sockets',
    )

    ip_address = fields.Field(
        column_name='IP Address',
        attribute='ip_address',
    )

    snmp_community = fields.Field(
        column_name='SNMP String',
        attribute='snmp_community',
    )

    assettag = fields.Field(
        column_name='Asset Tag',
        attribute='assettag',
    )

    customer_id = fields.Field(
        column_name='customer_id',
        attribute='customer_id',
    )

    status_column = 11

    class Meta:
        model = PDU
        id_field_index = 0
        id_field_name = 'name'
        customer_feild = 'customer_id'


class CloudResource(ErrorLoggingModelResource):
    name = fields.Field(
        column_name='Name',
        attribute='name',
    )

    customer_id = fields.Field(
        column_name='customer_id',
        attribute='customer_id',
    )

    class Meta:
        model = Cloud


class PrivateCloudResource(ErrorLoggingModelResource):
    name = fields.Field(
        column_name='Name',
        attribute='name',
    )

    customer_id = fields.Field(
        column_name='customer_id',
        attribute='customer_id',
    )

    platform_type = fields.Field(
        column_name='platform_type',
        attribute='platform_type',
    )

    storage = fields.Field(
        column_name='Storage',
        attribute='storage',
    )

    memory = fields.Field(
        column_name='Memory',
        attribute='memory',
    )

    colocation_cloud = fields.Field(
        column_name='Datacenter',
        attribute='colocation_cloud',
        widget=ForeignKeyWidget(ColoCloud, 'name')
    )

    status_column = 6

    class Meta:
        model = PrivateCloud
        id_field_index = 1
        id_field_name = 'name'
        customer_feild = 'customer_id'


class VMwareVcenterResource(ErrorLoggingModelResource):
    hostname = fields.Field(
        column_name='vCenter SSL URL',
        attribute='hostname',
        widget=VcenterURLWidget()
    )

    username = fields.Field(
        column_name='Username',
        attribute='username',
    )

    private_cloud = fields.Field(
        column_name='Name',
        attribute='private_cloud',
        widget=ForeignKeyWidget(PrivateCloud, 'name')
    )

    status_column = 6

    class Meta:
        model = VMwareVcenter
        id_field_index = 1
        id_field_name = 'private_cloud__name'
        customer_feild = 'private_cloud__customer_id'


class FirewallResource(ErrorLoggingModelResource):
    name = fields.Field(
        column_name='Name',
        attribute='name',
    )

    cabinet = fields.Field(
        column_name='Cabinet',
        attribute='cabinet',
        widget=ForeignKeyWidget(Cabinet, 'name')
    )

    manufacturer = fields.Field(
        column_name='Make',
        attribute='manufacturer',
        widget=ForeignKeyWidget(Manufacturer, 'name')
    )

    model = fields.Field(
        column_name='Model',
        attribute='model',
        widget=ManufacturerWidgetWithCreation(FirewallModel, 'name')
    )
    model = fields.Field(
        column_name='Asset Tag',
        attribute='asset_tag',
    )
    cloud_set = fields.Field(
        column_name='Cloud',
        attribute="cloud_set",
        widget=ManyToManyWidget(Cloud, field='name')
    )

    web_url = fields.Field(
        column_name='Web URL',
    )

    management_ip = fields.Field(
        column_name='Management IP',
        attribute='management_ip',
    )

    # ip_address = fields.Field(
    #     column_name='IP Address',
    #     attribute='ip_address',
    # )

    # snmp_community = fields.Field(
    #     column_name='SNMP String',
    #     attribute='snmp_community',
    # )

    size = fields.Field(
        column_name='Size',
        attribute='size',
    )

    position = fields.Field(
        column_name='Position',
        attribute='position',
        default=0,
    )

    customers = fields.Field(
        column_name='customers',
        attribute="customers",
        widget=ManyToManyWidget(Organization, 'id')
    )

    status_column = 14

    class Meta:
        model = Firewall
        id_field_index = 0
        id_field_name = 'name'
        customer_feild = 'customers'


class SwithcResource(ErrorLoggingModelResource):
    name = fields.Field(
        column_name='Name',
        attribute='name',
    )

    cabinet = fields.Field(
        column_name='Cabinet',
        attribute='cabinet',
        widget=ForeignKeyWidget(Cabinet, 'name')
    )

    manufacturer = fields.Field(
        column_name='Make',
        attribute='manufacturer',
        widget=ForeignKeyWidget(Manufacturer, 'name')
    )

    model = fields.Field(
        column_name='Model',
        attribute='model',
        widget=ManufacturerWidgetWithCreation(SwitchModel, 'name')
    )

    cloud_set = fields.Field(
        column_name='Cloud',
        attribute="cloud_set",
        widget=ManyToManyWidget(Cloud, field='name')
    )

    management_ip = fields.Field(
        column_name='Management IP',
        attribute='management_ip',
    )

    ip_address = fields.Field(
        column_name='IP Address',
        attribute='ip_address',
    )

    size = fields.Field(
        column_name='Size',
        attribute='size',
    )

    position = fields.Field(
        column_name='Position',
        attribute='position',
        default=0,
    )

    snmp_community = fields.Field(
        column_name='SNMP String',
        attribute='snmp_community',
    )

    customers = fields.Field(
        column_name='customers',
        attribute="customers",
        widget=ManyToManyWidget(Organization, 'id')
    )

    status_column = 14

    class Meta:
        model = Switch
        id_field_index = 0
        id_field_name = 'name'
        customer_feild = 'customers'


class LoadbalancerResource(ErrorLoggingModelResource):
    name = fields.Field(
        column_name='Name',
        attribute='name',
    )

    cabinet = fields.Field(
        column_name='Cabinet',
        attribute='cabinet',
        widget=ForeignKeyWidget(Cabinet, 'name')
    )

    manufacturer = fields.Field(
        column_name='Make',
        attribute='manufacturer',
        widget=ForeignKeyWidget(Manufacturer, 'name')
    )

    model = fields.Field(
        column_name='Model',
        attribute='model',
        widget=ManufacturerWidgetWithCreation(LoadBalancerModel, 'name')
    )

    cloud_set = fields.Field(
        column_name='Cloud',
        attribute="cloud_set",
        widget=ManyToManyWidget(Cloud, field='name')
    )

    management_ip = fields.Field(
        column_name='Management IP',
        attribute='management_ip',
    )

    ip_address = fields.Field(
        column_name='IP Address',
        attribute='ip_address',
    )

    snmp_community = fields.Field(
        column_name='SNMP String',
        attribute='snmp_community',
    )

    size = fields.Field(
        column_name='Size',
        attribute='size',
    )

    position = fields.Field(
        column_name='Position',
        attribute='position',
        default=0,
    )

    customers = fields.Field(
        column_name='customers',
        attribute="customers",
        widget=ManyToManyWidget(Organization, 'id')
    )

    status_column = 14

    class Meta:
        model = LoadBalancer
        id_field_index = 0
        id_field_name = 'name'
        customer_feild = 'customers'


class ServerResource(ErrorLoggingModelResource):
    name = fields.Field(
        column_name='Name',
        attribute='name',
    )

    cabinet = fields.Field(
        column_name='Cabinet',
        attribute='cabinet',
        widget=ForeignKeyWidget(Cabinet, 'name')
    )

    capacity_gb = fields.Field(
        column_name='Storage',
        attribute='capacity_gb',
    )

    num_cpus = fields.Field(
        column_name='CPUs',
        attribute='num_cpus',
    )

    memory_mb = fields.Field(
        column_name='Memory',
        attribute='memory_mb',
    )

    size = fields.Field(
        column_name='Size',
        attribute='size',
    )

    position = fields.Field(
        column_name='Position',
        attribute='position',
        default=0,
    )

    private_cloud = fields.Field(
        column_name='Cloud',
        attribute='private_cloud',
        widget=ForeignKeyWidget(PrivateCloud, 'name')
    )

    customer_id = fields.Field(
        column_name='customer_id',
        attribute='customer_id',
    )

    management_ip = fields.Field(
        column_name='Client/Machine SSH/RDP IP',
        attribute='management_ip',
    )

    ip_address = fields.Field(
        column_name='IP Address',
        attribute='ip_address',
    )

    snmp_community = fields.Field(
        column_name='SNMP String',
        attribute='snmp_community',
    )

    status_column = 18

    class Meta:
        model = Server
        id_field_index = 0
        id_field_name = 'name'
        customer_feild = 'customer_id'


class BMServerResource(ServerResource):
    status_column = 24


class HypervisorResource(ErrorLoggingModelResource):
    name = fields.Field(
        column_name='Name',
        attribute='name',
    )

    system = fields.Field(
        column_name='Name',
        attribute='system',
        widget=ForeignKeyWidget(Server, 'name')
    )

    os = fields.Field(
        column_name='Operating System',
        attribute='os',
        widget=OperatingSystemForeignKeyWidget(OperatingSystem, 'name')
    )

    virtualization_type = fields.Field(
        column_name='Virtualization Type',
        attribute='virtualization_type'
    )

    management_ip = fields.Field(
        column_name='Management IP',
        attribute='management_ip',
    )

    status_column = 20

    class Meta:
        model = Instance
        id_field_index = 0
        id_field_name = 'name'
        customer_feild = 'system__customer_id'


class BareMetalResource(ErrorLoggingModelResource):
    management_ip = fields.Field(
        column_name='Management IP',
        attribute='management_ip',
    )

    bmc_type = fields.Field(
        column_name='BM Type',
        attribute='bmc_type',
    )

    server = fields.Field(
        column_name='Name',
        attribute='server',
        widget=ForeignKeyWidget(Server, 'name')
    )

    os = fields.Field(
        column_name='Operating System',
        attribute='os',
        widget=OperatingSystemForeignKeyWidget(OperatingSystem, 'name')
    )

    status_column = 24

    class Meta:
        model = BMServer
        id_field_index = 0
        id_field_name = 'server__name'
        customer_feild = 'server__customer_id'


class IPMIControllerResource(ErrorLoggingModelResource):
    bm_server = fields.Field(
        column_name='Client/Machine SSH/RDP IP',
        attribute='bm_server',
        widget=ForeignKeyWidget(BMServer, 'management_ip')
    )

    ip = fields.Field(
        column_name='IP',
        attribute='ip',
    )

    username = fields.Field(
        column_name='Username',
        attribute='username',
    )

    password = fields.Field(
        column_name='Password',
        attribute='password',
    )
    status_column = 24

    class Meta:
        model = IPMIController
        id_field_index = 7
        id_field_name = 'ip'


class DRACControllerResource(ErrorLoggingModelResource):
    bm_server = fields.Field(
        column_name='Name',
        attribute='bm_server',
        widget=ForeignKeyWidget(BMServer, 'name')
    )

    ip = fields.Field(
        column_name='IP',
        attribute='ip',
    )

    username = fields.Field(
        column_name='Username',
        attribute='username',
    )

    password = fields.Field(
        column_name='Password',
        attribute='password',
    )

    status_column = 24

    class Meta:
        model = DRACController
        id_field_index = 7
        id_field_name = 'ip'


class VMWareVMResource(resources.ModelResource):
    id = fields.Field(
        column_name='ID',
        attribute='id',
        readonly=True
    )

    uuid = fields.Field(
        column_name='UUID',
        attribute='uuid',
        readonly=True
    )

    name = fields.Field(
        column_name='Name',
        attribute='name',
        readonly=True
    )

    cloud_name = fields.Field(
        column_name='Cloud',
        attribute='cloud__name',
        readonly=True
    )

    management_ip = fields.Field(
        column_name='Management IP',
        attribute='management_ip',
    )

    ip_address = fields.Field(
        column_name='IP Address',
        attribute='ip_address',
    )

    snmp_community = fields.Field(
        column_name='SNMP String',
        attribute='snmp_community',
    )

    monitoring_status = fields.Field(
        column_name='Monitoring Status',
        attribute='monitoring_status',
    )

    def dehydrate_monitoring_status(self, vm):
        if hasattr(vm, 'observium_vmware_vm'):
            return "Enabled"
        return "Not Enabled"

    class Meta:
        model = VmwareVmMigration
        fields = (
            'id', 'uuid', 'name', 'cloud_name', 'management_ip', 'ip_address', 'snmp_community', 'monitoring_status'
        )
        export_order = (
            'id', 'uuid', 'name', 'cloud_name', 'management_ip', 'ip_address', 'snmp_community', 'monitoring_status'
        )


class OpenStackVMResource(resources.ModelResource):
    id = fields.Field(
        column_name='ID',
        attribute='id',
        readonly=True
    )

    uuid = fields.Field(
        column_name='UUID',
        attribute='uuid',
        readonly=True
    )

    name = fields.Field(
        column_name='Name',
        attribute='name',
        readonly=True
    )

    cloud_name = fields.Field(
        column_name='Cloud',
        attribute='cloud__name',
        readonly=True
    )

    management_ip = fields.Field(
        column_name='Management IP',
        attribute='management_ip',
    )

    ip_address = fields.Field(
        column_name='IP Address',
        attribute='ip_address',
    )

    snmp_community = fields.Field(
        column_name='SNMP String',
        attribute='snmp_community',
    )

    monitoring_status = fields.Field(
        column_name='Monitoring Status',
        attribute='monitoring_status',
    )

    def dehydrate_monitoring_status(self, vm):
        if hasattr(vm, 'observium_openstack_vm'):
            return "Enabled"
        return "Not Enabled"

    class Meta:
        model = OpenStackVmMigration
        fields = (
            'id', 'uuid', 'name', 'cloud_name', 'management_ip', 'ip_address', 'snmp_community', 'monitoring_status'
        )
        export_order = (
            'id', 'uuid', 'name', 'cloud_name', 'management_ip', 'ip_address', 'snmp_community', 'monitoring_status'
        )


class VCloudVMResource(resources.ModelResource):
    id = fields.Field(
        column_name='ID',
        attribute='id',
        readonly=True
    )

    uuid = fields.Field(
        column_name='UUID',
        attribute='uuid',
        readonly=True
    )

    name = fields.Field(
        column_name='Name',
        attribute='name',
        readonly=True
    )

    cloud_name = fields.Field(
        column_name='Cloud',
        attribute='cloud__name',
        readonly=True
    )

    management_ip = fields.Field(
        column_name='Management IP',
        attribute='management_ip',
    )

    ip_address = fields.Field(
        column_name='IP Address',
        attribute='ip_address',
    )

    snmp_community = fields.Field(
        column_name='SNMP String',
        attribute='snmp_community',
    )

    monitoring_status = fields.Field(
        column_name='Monitoring Status',
        attribute='monitoring_status',
    )

    def dehydrate_monitoring_status(self, vm):
        if hasattr(vm, 'observium_vcloud_vm'):
            return "Enabled"
        return "Not Enabled"

    class Meta:
        model = VCloudVirtualMachines
        fields = (
            'id', 'uuid', 'name', 'cloud_name', 'management_ip', 'ip_address', 'snmp_community', 'monitoring_status'
        )
        export_order = (
            'id', 'uuid', 'name', 'cloud_name', 'management_ip', 'ip_address', 'snmp_community', 'monitoring_status'
        )


class ProxmoxVMResource(resources.ModelResource):
    id = fields.Field(
        column_name='ID',
        attribute='id',
        readonly=True
    )

    uuid = fields.Field(
        column_name='UUID',
        attribute='uuid',
        readonly=True
    )

    name = fields.Field(
        column_name='Name',
        attribute='vm_name',
        readonly=True
    )

    cloud_name = fields.Field(
        column_name='Cloud',
        attribute='cluster__private_cloud__name',
        readonly=True
    )

    management_ip = fields.Field(
        column_name='Management IP',
        attribute='management_ip',
    )

    ip_address = fields.Field(
        column_name='IP Address',
        attribute='ip_address',
    )

    snmp_community = fields.Field(
        column_name='SNMP String',
        attribute='snmp_community',
    )

    monitoring_status = fields.Field(
        column_name='Monitoring Status',
        attribute='monitoring_status',
    )

    def dehydrate_monitoring_status(self, vm):
        if hasattr(vm, 'observium_vcloud_vm'):
            return "Enabled"
        return "Not Enabled"

    class Meta:
        model = ProxmoxVM
        fields = (
            'id', 'uuid', 'name', 'cloud_name', 'management_ip', 'ip_address', 'snmp_community', 'monitoring_status'
        )
        export_order = (
            'id', 'uuid', 'name', 'cloud_name', 'management_ip', 'ip_address', 'snmp_community', 'monitoring_status'
        )

Model_Dict = {
    'Cabinets': Cabinet,
    'Datacenters': ColoCloud,
    'PDUs': PDU,
    'Firewalls': Firewall,
    'Load_Balancers': LoadBalancer,
    'Switches': Switch,
    'Hypervisors': Server,
    'Bare_Metals': BMServer,
    'Storage': StorageDevice,
    'MAC_Mini': MacDevice,
    'Mobile_Devices': MobileDevice,
    'Databases': DatabaseServer
}


def validate_onbaording_data(data, customer):
    for key in data.keys():
        if key == "Bare_Metals":
            for value in data[key]:
                if value:
                    query_data = BMServer.objects.filter(server__customer=customer).values_list('server__name', flat=True)
                    value['onboarding_status'] = "Onboarded" if value['name'] in query_data else None

        elif key == "Mobile_Devices":
            for value in data[key]:
                if value:
                    query_data = MobileDevice.objects.filter(customer=customer).values_list('serial_number', flat=True)
                    value['onboarding_status'] = "Onboarded" if str(value['serial_number']) in query_data else None

        elif key == "Databases":
            for value in data[key]:
                if value:
                    bm_set = BMServer.objects.filter(server__customer=customer)
                    vm_set = VirtualMachine.objects.filter(customer=customer)
                    vmware_set = VmwareVmMigration.objects.filter(cloud__customer=customer)
                    vcvm_set = VCloudVirtualMachines.objects.filter(cloud__customer=customer)
                    kvm_set = ProxmoxVM.objects.filter(cluster__private_cloud__customer=customer)
                    hyp_set = HypervVM.objects.filter(cluster__private_cloud__customer=customer)
                    opvm_set = OpenStackVmMigration.objects.filter(cloud__customer=customer)
                    
                    value['onboarding_status'] = None

                    for bm in bm_set:
                        dbs = bm.database_server.all()
                        if dbs:
                            db_name = dbs.first().db_instance_name
                        else:
                            db_name = None
                        if value['management_ip'] == bm.management_ip and value['name'] == db_name:
                            value['onboarding_status'] = "Onboarded"
                            break

                    # if not value['onboarding_status']:
                    #     for vm in vm_set:
                    #         dbs = vm.database_server.all()
                    #         if dbs:
                    #             db_name = dbs.first().db_instance_name
                    #         else:
                    #             db_name = None
                    #         if value['management_ip'] == vm.management_ip and value['name'] == db_name:
                    #             value['onboarding_status'] = "Onboarded"
                    #             break

                    for vm in vm_set:
                        dbs = vm.database_server.all()
                        if dbs:
                            db_name = dbs.first().db_instance_name
                        else:
                            db_name = None
                        if value['management_ip'] == vm.management_ip and value['name'] == db_name:
                            value['onboarding_status'] = "Onboarded"
                            break

                    for vmw in vmware_set:
                        dbs = vmw.database_server.all()
                        if dbs:
                            db_name = dbs.first().db_instance_name
                        else:
                            db_name = None
                        if value['management_ip'] == vmw.management_ip and value['name'] == db_name:
                            value['onboarding_status'] = "Onboarded"
                            break
                        
                    for vc in vcvm_set:
                        dbs = vc.database_server.all()
                        if dbs:
                            db_name = dbs.first().db_instance_name
                        else:
                            db_name = None
                        if value['management_ip'] == vc.management_ip and value['name'] == db_name:
                            value['onboarding_status'] = "Onboarded"
                            break
                        
                    for kvm in kvm_set:
                        dbs = kvm.database_server.all()
                        if dbs:
                            db_name = dbs.first().db_instance_name
                        else:
                            db_name = None
                        if value['management_ip'] == kvm.management_ip and value['name'] == db_name:
                            value['onboarding_status'] = "Onboarded"
                            break
                        
                    for hyp in hyp_set:
                        dbs = hyp.database_server.all()
                        if dbs:
                            db_name = dbs.first().db_instance_name
                        else:
                            db_name = None
                        if value['management_ip'] == hyp.management_ip and value['name'] == db_name:
                            value['onboarding_status'] = "Onboarded"
                            break
                        
                    for op in opvm_set:
                        dbs = op.database_server.all()
                        if dbs:
                            db_name = dbs.first().db_instance_name
                        else:
                            db_name = None
                        if value['management_ip'] == op.management_ip and value['name'] == db_name:
                            value['onboarding_status'] = "Onboarded"
                            break
                    
        elif key in ["Datacenters", "Hypervisors", "Storage", "MAC_Mini", "PDUs"]:
            for value in data[key]:
                if value:
                    query_data = Model_Dict[key].objects.filter(customer=customer).values_list('name', flat=True)
                    value['onboarding_status'] = "Onboarded" if value['name'] in query_data else None

        elif key in ["Cabinets", "Firewalls", "Load_Balancers", "Switches"]:
            for value in data[key]:
                if value:
                    query_data = Model_Dict[key].objects.filter(customers__in=[customer]).values_list('name', flat=True)
                    value['onboarding_status'] = "Onboarded" if value['name'] in query_data else None
    return data


def prepare_json_data(data, device_type, unique_id):
    json_dict = {}
    if device_type == "Datacenters":
        json_dict = {
            'name': data.name,
            'location': data.location,
            'onboarding_status': 'Onboarded',
            'db_id': data.id,
            'unique_id': unique_id
        }
    elif device_type == "Cabinets":
        json_dict = {
            "datacenter": data.colocloud_set.all()[0].name,
            "db_id": data.id,
            "name": data.name,
            "renewal": data.renewal,
            "contract_end_date": data.contract_end_date.strftime("%m/%d/%Y, %H:%M:%S") if data.contract_end_date else None,
            "cost": data.cost,
            "contract_start_date": data.contract_start_date.strftime("%m/%d/%Y, %H:%M:%S") if data.contract_start_date else None,
            "annual_escalation": data.annual_escalation,
            "model": data.model,
            "onboarding_status": "Onboarded",
            "size": data.size,
            "unique_id": unique_id
        }
    elif device_type == "Databases":
        json_dict = {
            "name": data.db_instance_name,
            "db_type": data.db_type.name,
            "port": data.port,
            'onboarding_status': 'Onboarded',
            "management_ip": data.management_ip,
            "unique_id": unique_id
        }
    elif device_type == "PDUs":
        # power_circuit = data.power_circuit.name + " (" + data.power_circuit.\
        #     voltagetype.voltage_type + "/" + data.power_circuit.ampstype.\
        #     amps_type + ")"
        json_dict = {
            # "power_circuit": power_circuit,
            "power_circuit": str(data.power_circuit.name),
            "db_id": data.id,
            "pdu_type": data.pdu_type,
            "name": data.name,
            "manufacturer": data.manufacturer.name,
            "model": data.model.model_number,
            "onboarding_status": "Onboarded",
            "cabinet": data.cabinet.name if data.cabinet else None,
            "asset_tag": data.assettag,
            "position": data.position,
            "ip_address": data.ip_address,
            "sockets": data.sockets,
            "size": data.size,
            "unique_id": unique_id,
            "snmp_ip": data.ip_address,
            "collector": data.collector.name
        }
    elif device_type == "Firewalls":
        json_dict = {
            "db_id": data.id,
            "name": data.name,
            "cabinet": data.cabinet.name if data.cabinet else None,
            "asset_tag": data.asset_tag,
            "management_ip": data.management_ip,
            "position": data.position,
            "model": data.model.name,
            "manufacturer": data.model.manufacturer.name,
            "onboarding_status": "Onboarded",
            "cloud": [item.name for item in data.cloud_set.all()],
            "size": data.size,
            "unique_id": unique_id,
            "snmp_ip": data.ip_address,
            "datacenter": data.datacenter.name,
            "collector": data.collector.name
        }
    elif device_type == "Load_Balancers":
        json_dict = {
            "db_id": data.id,
            "name": data.name,
            "cabinet": data.cabinet.name if data.cabinet else None,
            "asset_tag": data.asset_tag,
            "management_ip": data.management_ip,
            "position": data.position,
            "model": data.model.name,
            "manufacturer": data.model.manufacturer.name,
            "onboarding_status": "Onboarded",
            "cloud": [item.name for item in data.cloud_set.all()],
            "size": data.size,
            "unique_id": unique_id,
            "snmp_ip": data.ip_address,
            "datacenter": data.datacenter.name,
            "collector": data.collector.name
        }
    elif device_type == "Switches":
        json_dict = {
            "db_id": data.id,
            "name": data.name,
            "cabinet": data.cabinet.name if data.cabinet else None,
            "asset_tag": data.asset_tag,
            "management_ip": data.management_ip,
            "position": data.position,
            "model": data.model.name,
            "manufacturer": data.model.manufacturer.name,
            "onboarding_status": "Onboarded",
            "cloud": [item.name for item in data.cloud_set.all()],
            "size": data.size,
            "unique_id": unique_id,
            "snmp_ip": data.ip_address,
            "datacenter": data.datacenter.name,
            "collector": data.collector.name
        }
    elif device_type == "Hypervisors":
        json_dict = {
            "memory_mb": data.memory_mb,
            "db_id": data.id,
            "name": data.name,
            "num_cpus": data.num_cpus,
            "capacity_gb": data.capacity_gb,
            "num_cores": data.num_cores,
            "virtualization_type": data.instance.virtualization_type,
            "onboarding_status": "Onboarded",
            "cabinet": data.cabinet.name if data.cabinet else None,
            "asset_tag": data.asset_tag,
            "management_ip": data.management_ip,
            "position": data.position,
            "model": data.model.name,
            "manufacturer": data.manufacturer.name,
            "os": data.os.full_name if data.os else None,
            "private_cloud": data.private_cloud.name if data.private_cloud else None,
            "size": data.size,
            "unique_id": unique_id,
            "snmp_ip": data.ip_address,
            "datacenter": data.datacenter.name,
            "collector": data.collector.name
        }
    elif device_type == "Bare_Metals":
        if isinstance(data, Server):
            data = BMServer.objects.get(server=data)
        json_dict = {
            "memory_mb": data.server.memory_mb,
            "name": data.server.name,
            "num_cpus": data.server.num_cpus,
            "capacity_gb": data.server.capacity_gb,
            "num_cores": data.server.num_cores,
            "cabinet": data.server.cabinet.name if data.server.cabinet else None,
            "asset_tag": data.server.asset_tag,
            "management_ip": data.management_ip,
            "position": data.server.position,
            "bmc_type": data.bmc_type,
            "model": data.server.model.name,
            "manufacturer": data.server.manufacturer.name,
            "os": data.os.full_name if data.os else None,
            "private_cloud": data.server.private_cloud.name if data.server.private_cloud else None,
            "size": data.server.size,
            "onboarding_status": "Onboarded",
            "db_id": data.id,
            "unique_id": unique_id,
            "snmp_ip": data.server.ip_address,
            "datacenter": data.datacenter.name,
            "collector": data.collector.name
        }
    elif device_type == "Storage":
        json_dict = {
            "name": data.name,
            "cabinet": data.cabinet.name if data.cabinet else None,
            "asset_tag": data.asset_tag,
            "management_ip": data.management_ip,
            "position": data.position,
            "model": data.model.name,
            "manufacturer": data.manufacturer.name,
            "os": data.os.full_name if data.os else None,
            "private_cloud": data.private_cloud.name if data.private_cloud else None,
            "size": data.size,
            'onboarding_status': 'Onboarded',
            'db_id': data.id,
            'unique_id': unique_id,
            "snmp_ip": data.ip_address,
            "datacenter": data.datacenter.name,
            "collector": data.collector.name
        }
    elif device_type == "MAC_Mini":
        json_dict = {
            "memory_mb": data.memory_mb,
            "db_id": data.id,
            "name": data.name,
            "num_cpus": data.num_cpus,
            "capacity_gb": data.capacity_gb,
            "num_cores": data.num_cores,
            "onboarding_status": "Onboarded",
            "cabinet": data.cabinet.name if data.cabinet else None,
            "asset_tag": data.asset_tag,
            "management_ip": data.management_ip,
            "serial_number": data.serial_number,
            "model": data.model.name,
            "os": data.os.full_name if data.os else None,
            "private_cloud": data.private_cloud.name if data.private_cloud else None,
            "manufacturer": data.manufacturer.name,
            "unique_id": unique_id,
            "snmp_ip": data.ip_address,
            "datacenter": data.datacenter.name,
            "collector": data.collector.name
        }
    elif device_type == "Mobile_Devices":
        json_dict = {
            "tagged_device": data.tagged_device.name if data.tagged_device else None,
            "db_id": data.id,
            "name": data.name,
            "platform": data.platform,
            "device_type": data.device_type,
            "serial_number": data.serial_number,
            "model": data.model,
            "ip_address": data.ip_address,
            "onboarding_status": "Onboarded",
            "unique_id": unique_id,
            "datacenter": data.datacenter.name if data.datacenter else None,
            "collector": data.collector.name
            # "snmp_ip": data.snmp_ip
        }
    return json_dict


def prepare_json_temp(data, device_type, unique_id):
    json_dict = {}
    try:
        if device_type == "Datacenters":
            json_dict = {
                'name': data['name'],
                'location': data['location'],
                'onboarding_status': None,
                'unique_id': unique_id
            }
        elif device_type == "Cabinets":
            json_dict = {
                "datacenter": ColoCloud.objects.get(uuid=data['datacenter']).name if data['datacenter'] else None,
                "name": data['name'],
                "renewal": data['renewal'],
                "contract_end_date": data['contract_end_date'],
                "cost": data['cost'],
                "contract_start_date": data['contract_start_date'],
                "annual_escalation": data['annual_escalation'],
                "model": data['model'],
                "onboarding_status": None,
                "size": data['size'],
                "unique_id": unique_id
            }
        elif device_type == "Databases":
            json_dict = {
                "name": data["name"],
                "db_type": data["db_type"]["name"] if type(data["db_type"]) == dict else data["db_type"],
                "port": data["port"],
                "management_ip": data["management_ip"],
                "unique_id": unique_id
            }
        elif device_type == "PDUs":
            json_dict = {
                "pdu_type": data['pdu_type'],
                "name": data['name'],
                "manufacturer": PDUManufacturer.objects.get(id=int(data['manufacturer']['id'])).name if data['manufacturer']['id'] else None,
                "model": PDUModel.objects.get(id=int(data['model']['id'])).model_number if data['model']['id'] else None,
                "onboarding_status": None,
                "cabinet": Cabinet.objects.get(id=int(data['cabinet']['id'])).name if data['cabinet']['id'] else None,
                "asset_tag": data['asset_tag'],
                "position": data['position'],
                "management_ip": data.get('management_ip'),
                "sockets": data['sockets'],
                "size": data['size'],
                "power_circuit": PowerCircuit.objects.get(id=int(data['power_circuit']['id'])).name if data['power_circuit']['id'] else None,
                "unique_id": unique_id,
                "snmp_ip": data['snmp_ip'],
                "collector": AgentConfig.objects.get(uuid=data['collector']['uuid']).name if data['collector'][
                    'uuid'] else None
            }
        elif device_type == "Firewalls":
            json_dict = {
                "name": data['name'],
                "cabinet": Cabinet.objects.get(id=int(data['cabinet']['id'])).name if data['cabinet']['id'] else None,
                "asset_tag": data['asset_tag'],
                "management_ip": data['management_ip'],
                "position": data['position'],
                "model": FirewallModel.objects.get(id=int(data['model']['id'])).name if data['model']['id'] else None,
                "manufacturer": Manufacturer.objects.get(id=int(data['manufacturer'])).name if data['manufacturer'] else None,
                "onboarding_status": None,
                "cloud": [item['name'] for item in data['cloud']],
                "size": data['size'],
                "unique_id": unique_id,
                "snmp_ip": data['snmp_ip'],
                "datacenter": ColoCloud.objects.get(uuid=data['datacenter']['uuid']).name if data['datacenter']['uuid'] else None,
                "collector": AgentConfig.objects.get(uuid=data['collector']['uuid']).name if data['collector'][
                    'uuid'] else None
            }
        elif device_type == "Load_Balancers":
            json_dict = {
                "name": data['name'],
                "cabinet": Cabinet.objects.get(id=int(data['cabinet']['id'])).name if data['cabinet']['id'] else None,
                "asset_tag": data['asset_tag'],
                "management_ip": data['management_ip'],
                "position": data['position'],
                "model": LoadBalancerModel.objects.get(id=int(data['model']['id'])).name if data['model']['id'] else None,
                "manufacturer": Manufacturer.objects.get(id=int(data['manufacturer'])).name if data['manufacturer'] else None,
                "onboarding_status": None,
                "cloud": [item['name'] for item in data['cloud']],
                "size": data['size'],
                "unique_id": unique_id,
                "snmp_ip": data['snmp_ip'],
                "datacenter": ColoCloud.objects.get(uuid=data['datacenter']['uuid']).name if data['datacenter']['uuid'] else None,
                "collector": AgentConfig.objects.get(uuid=data['collector']['uuid']).name if data['collector'][
                    'uuid'] else None
            }
        elif device_type == "Switches":
            json_dict = {
                "name": data['name'],
                "cabinet": Cabinet.objects.get(id=int(data['cabinet']['id'])).name if data['cabinet']['id'] else None,
                "asset_tag": data['asset_tag'],
                "management_ip": data['management_ip'],
                "position": data['position'],
                "model": SwitchModel.objects.get(id=int(data['model']['id'])).name if data['model']['id'] else None,
                "manufacturer": Manufacturer.objects.get(id=int(data['manufacturer'])).name if data['manufacturer'] else None,
                "onboarding_status": None,
                "cloud": [item['name'] for item in data['cloud']],
                "size": data['size'],
                "unique_id": unique_id,
                "snmp_ip": data['snmp_ip'],
                "datacenter": ColoCloud.objects.get(uuid=data['datacenter']['uuid']).name if data['datacenter']['uuid'] else None,
                "collector": AgentConfig.objects.get(uuid=data['collector']['uuid']).name if data['collector'][
                    'uuid'] else None
            }
        elif device_type == "Hypervisors":
            json_dict = {
                "memory_mb": data['memory_mb'],
                "name": data['name'],
                "num_cpus": data['num_cpus'],
                "capacity_gb": data['capacity_gb'],
                "num_cores": data['num_cores'],
                "virtualization_type": data['virtualization_type'],
                "onboarding_status": None,
                "cabinet": Cabinet.objects.get(id=int(data['cabinet']['id'])).name if data['cabinet']['id'] else None,
                "asset_tag": data['asset_tag'],
                "management_ip": data['management_ip'],
                "position": data['position'],
                "model": ServerModel.objects.get(id=int(data['model']['id'])).name if data['model']['id'] else None,
                "manufacturer": ServerManufacturer.objects.get(id=int(data['manufacturer']['id'])).name if data['manufacturer']['id'] else None,
                "os": OperatingSystem.objects.get(id=int(data['os']['id'])).name if data['os']['id'] else None,
                "private_cloud": PrivateCloud.objects.get(id=int(data['private_cloud']['id'])).name if data['private_cloud']['id'] else None,
                "size": data['size'],
                "unique_id": unique_id,
                "snmp_ip": data['snmp_ip'],
                "datacenter": ColoCloud.objects.get(uuid=data['datacenter']['uuid']).name if data['datacenter']['uuid'] else None,
                "collector": AgentConfig.objects.get(uuid=data['collector']['uuid']).name if data['collector'][
                    'uuid'] else None
            }
        elif device_type == "Bare_Metals":
            json_dict = {
                "memory_mb": data['memory_mb'],
                "name": data['name'],
                "num_cpus": data['num_cpus'],
                "capacity_gb": data['capacity_gb'],
                "num_cores": data['num_cores'],
                "cabinet": Cabinet.objects.get(id=int(data['cabinet']['id'])).name if data['cabinet']['id'] else None,
                "asset_tag": data['asset_tag'],
                "management_ip": data['management_ip'],
                "position": data['position'],
                "bmc_type": data['bmc_type'],
                "model": ServerModel.objects.get(id=int(data['model']['id'])).name if data['model']['id'] else None,
                "manufacturer": ServerManufacturer.objects.get(id=int(data['manufacturer']['id'])).name if data['manufacturer']['id'] else None,
                "os": OperatingSystem.objects.get(id=int(data['os']['id'])).name if data['os']['id'] else None,
                "private_cloud": PrivateCloud.objects.get(id=int(data['private_cloud']['id'])).name if data['private_cloud']['id'] else None,
                "size": data['size'],
                "onboarding_status": None,
                "unique_id": unique_id,
                "snmp_ip": data['snmp_ip'],
                "datacenter": ColoCloud.objects.get(uuid=data['datacenter']['uuid']).name if data['datacenter']['uuid'] else None,
                "collector": AgentConfig.objects.get(uuid=data['collector']['uuid']).name if data['collector'][
                    'uuid'] else None
            }
        elif device_type == "Storage":
            json_dict = {
                "name": data['name'],
                "cabinet": Cabinet.objects.get(id=int(data['cabinet']['id'])).name if data['cabinet']['id'] else None,
                "asset_tag": data['asset_tag'],
                "management_ip": data['management_ip'],
                "position": data['position'],
                "model": StorageModel.objects.get(id=int(data['model']['id'])).name if data['model']['id'] else None,
                "manufacturer": StorageManufacturer.objects.get(id=int(data['manufacturer']['id'])).name if data['manufacturer']['id'] else None,
                "os": OperatingSystem.objects.get(id=int(data['os']['id'])).name if data['os']['id'] else None,
                "private_cloud": PrivateCloud.objects.get(id=int(data['private_cloud']['id'])).name if data['private_cloud']['id'] else None,
                "size": data['size'],
                'onboarding_status': None,
                'unique_id': unique_id,
                "snmp_ip": data['snmp_ip'],
                "datacenter": ColoCloud.objects.get(uuid=data['datacenter']['uuid']).name if data['datacenter']['uuid'] else None,
                "collector": AgentConfig.objects.get(uuid=data['collector']['uuid']).name if data['collector'][
                    'uuid'] else None
            }
        elif device_type == "MAC_Mini":
            json_dict = {
                "memory_mb": data['memory_mb'],
                "name": data['name'],
                "num_cpus": data['num_cpus'],
                "capacity_gb": data['capacity_gb'],
                "num_cores": data['num_cores'],
                "onboarding_status": None,
                "cabinet": Cabinet.objects.get(id=int(data['cabinet']['id'])).name if data['cabinet']['id'] else None,
                "asset_tag": data['asset_tag'],
                "management_ip": data['management_ip'],
                "serial_number": data['serial_number'],
                "model": ServerModel.objects.get(id=int(data['model']['id'])).name if data['model']['id'] else None,
                "os": OperatingSystem.objects.get(id=int(data['os']['id'])).name if data['os']['id'] else None,
                "private_cloud": PrivateCloud.objects.get(id=int(data['private_cloud']['id'])).name if data['private_cloud']['id'] else None,
                "manufacturer": ServerManufacturer.objects.get(id=int(data['manufacturer']['id'])).name if data['manufacturer']['id'] else None,
                "unique_id": unique_id,
                "snmp_ip": data['snmp_ip'],
                "datacenter": ColoCloud.objects.get(uuid=data['datacenter']['uuid']).name if data['datacenter']['uuid'] else None,
                "collector": AgentConfig.objects.get(uuid=data['collector']['uuid']).name if data['collector'][
                    'uuid'] else None
            }
        elif device_type == "Mobile_Devices":
            if data["platform"] == 'ios':
                json_dict["tagged_device"] = MacDevice.objects.get(id=int(data['tagged_device'][0]['id'])).name if data.get('tagged_device') else None
            else:
                json_dict["tagged_device"] = BMServer.objects.get(server__id=int(data['tagged_device'][0]['id'])).name if data.get('tagged_device') else None
            json_dict = {
                "name": data['name'],
                "platform": data['platform'],
                "device_type": data['device_type'],
                "serial_number": data['serial_number'],
                "model": MobileDeviceModel.objects.get_or_create(name=data['model'])[0].name if data['model'] else None,
                "ip_address": data['ip_address'],
                "onboarding_status": None,
                "unique_id": unique_id,
                "datacenter": ColoCloud.objects.get(uuid=data['datacenter']['uuid']).name if 'datacenter' in data and 'uuid' in data['datacenter'] and data['datacenter']['uuid'] else None,
                "collector": AgentConfig.objects.get(uuid=data['collector']['uuid']).name if data['collector'][
                    'uuid'] else None
            }
    except Exception as e:
        return data, e

    return json_dict, None

# def insert_datasheet_data(model_query, column_num, field_name, row_count=3):
#     for item in model_query:
#         copy_sheet.cell(row=row_count, column=8).value = i['name']
#         row_count += 1
#     pass
