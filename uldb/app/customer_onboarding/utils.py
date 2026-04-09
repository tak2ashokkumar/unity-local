import base64
import csv
import openpyxl
import os
import uuid
from copy import deepcopy
from django.http import HttpResponse
from datetime import datetime
from django.contrib.contenttypes.models import ContentType

from agent.models import AgentConfig
from app.datacenter.models import Cabinet
from app.inventory.models import (
    Switch, Firewall, LoadBalancer, Server, BareMetal, BMServer,
    StorageDevice, MobileDevice, MacDevice, DatabaseServer,
    CustomAttribute, SwitchModel, FirewallModel, LoadBalancerModel,
    ServerModel, Manufacturer, ServerManufacturer, OperatingSystem,
    DatabaseEntity, StorageManufacturer, StorageModel, DatabaseType
)
from cloud.CloudService.models import ColoCloud, PrivateCloud
from cloud.VmwareAdapter.models import VmwareVmMigration

DEVICE_MODEL_MAP = {
    "switch": (Switch, "Switches"),
    "firewall": (Firewall, "Firewalls"),
    "load_balancer": (LoadBalancer, "Load Balancers"),
    "hypervisor": (Server, "Hypervisors"),
    "baremetal": (BMServer, "Bare Metals"),
    "storage": (StorageDevice, "Storage"),
    "mobile": (MobileDevice, "Mobile Devices"),
    "mac_device": (MacDevice, "MAC Mini"),
    "database_servers": (DatabaseServer, "Databases"),
    "database_entity": (DatabaseEntity, 'Database Entity'),
    "vmware": (VmwareVmMigration, 'VMWare VM')
}

NETWORK_DEVICE_HEADERS = [
    'name', 'management_ip', 'manufacturer',
    'model', 'datacenter', 'cloud',
    'cabinet', 'size', 'position',
    'asset_tag', 'snmp_ip', 'collector'
]
OTHER_COMMON_HEADERS = [
    'name', 'management_ip', 'manufacturer',
    'model', 'datacenter', 'private_cloud',
    'cabinet', 'size', 'position'
]
SERVER_HEADERS = ['num_cpus', 'num_cores', 'memory_mb', 'capacity_gb']
HEADERS_DICT = {
    "Datacenters": [
        'name', 'location'
    ],
    "Cabinets": [
        'name', 'size', 'datacenter',
        'model', 'contract_start_date',
        'contract_end_date', 'cost',
        'annual_escalation', 'renewal'
    ],
    "PDUs": [
        'name', 'cabinet', 'size',
        'power_circuit', 'pdu_type',
        'position', 'manufacturer', 'model',
        'ip_address', 'asset_tag', 'snmp_ip',
        'sockets', 'collector'
    ],
    "Firewalls": NETWORK_DEVICE_HEADERS,
    "Switches": NETWORK_DEVICE_HEADERS,
    "Load Balancers": NETWORK_DEVICE_HEADERS,
    "Hypervisors": OTHER_COMMON_HEADERS + SERVER_HEADERS + [
        'virtualization_type', 'os', 'asset_tag',
        'snmp_ip',  'collector'
    ],
    "Bare Metals": OTHER_COMMON_HEADERS + SERVER_HEADERS + [
        'os', 'bmc_type', 'asset_tag',
        'snmp_ip',  'collector'
    ],
    "Storage": OTHER_COMMON_HEADERS + [
        'os', 'asset_tag', 'snmp_ip',  'collector'
    ],
    "Mobile Devices": [
        'name', 'serial_number', 'model',
        'datacenter', 'ip_address', 'platform',
        'device_type', 'tagged_device', 'collector'
    ],
    "MAC Mini": [
        'name', 'serial_number', 'manufacturer',
        'model', 'datacenter', 'private_cloud',
        'cabinet', 'os', 'management_ip',
        'num_cpus', 'num_cores', 'memory_mb',
        'capacity_gb', 'asset_tag', 'snmp_ip',
        'collector'
    ],
    "Databases": ['db_instance_name', 'db_type', 'port', 'management_ip'],
    "Database Entity": ['name', 'short_description'],
    "VMWare VMs": ['name']
}

DEVICE_MISMATCH_MAP = {
    "switch": "switch",
    "firewall": "firewall",
    "load_balancer": "loadbalancers",
    "hypervisor": "hypervisors",
    "baremetal": "bm-servers",
    "mac_device": "mac-mini",
    "storage": "storage-devices",
}

BASE_DIR = 'media/bulk_update/'

def validate_csv_value(value):
    """
    Convert Django model fields/objects into something openpyxl can write.
    Works in Python 2.7.
    """
    if value is None:
        return ""
    # Handle ForeignKeys / model instances
    if hasattr(value, '__unicode__'):
        return unicode(value)
    if hasattr(value, '__str__'):
        return str(value)
    # Numbers are fine
    if isinstance(value, (int, float, long, bool)):
        return value
    if hasattr(value, '__iter__') and not isinstance(value, (basestring, dict)):
        return u", ".join([validate_csv_value(v) for v in value])
    # Fallback
    return unicode(value)


def export_device_csv(device_type, uuids, customer):
    """
    Generate Excel file for devices and return the file path.
    """

    for k,v in DEVICE_MISMATCH_MAP.items():
        if device_type == v:
            device_type = k
    if device_type not in DEVICE_MODEL_MAP:
        raise ValueError("Invalid device type")

    model_class, sheet_name = DEVICE_MODEL_MAP[device_type]
    if uuids == 'all':
        if device_type in [ "switch", "firewall", "load_balancer"]:
            devices = model_class.objects.filter(customers=customer)
        elif device_type in ["baremetal"]:
            devices = model_class.objects.filter(server__customer=customer)
        elif device_type in ["storage", "mobile", "mac_device", "database_servers","hypervisor"]:
            devices = model_class.objects.filter(customer=customer)
        elif device_type in ["database_entity"]:
            devices = model_class.objects.filter(database_server__customer=customer)
        elif device_type in ["vmware"]:
            devices = model_class.objects.filter(cloud__customer=customer)
        else:
            raise ValueError("No device type found")
    else:
        devices = model_class.objects.filter(uuid__in=uuids)
    # if not devices.exists():
    #     raise ValueError("No devices found for given UUIDs")

    template_path = os.path.join(BASE_DIR, 'BulkUpdateDevices.xlsx')

    wb = openpyxl.load_workbook(template_path, data_only=True)
    wb = deepcopy(wb)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError("{} sheet not found in template".format(sheet_name))

    for s in wb.sheetnames:
        if s != sheet_name:
            del wb[s]

    sheet = wb[sheet_name]
    template_headers = [cell.value for cell in sheet[1] if cell.value]

    model_fields = HEADERS_DICT[sheet_name]

    ct = ContentType.objects.get_for_model(model_class)
    custom_attrs = CustomAttribute.objects.filter(content_type=ct, customer=customer)
    custom_headers = [str(attr.name) for attr in custom_attrs]

    headers_with_custom = template_headers + custom_headers + [u"UNITY_ID"]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = "{}_devices_{}.csv".format(device_type.lower(), timestamp)
    full_path = os.path.join(BASE_DIR, filename)

    with open(full_path, "wb") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers_with_custom)

        for device in devices:
            bm_device = device
            row = []
            custom_key = []
            if device.DEVICE_TYPE == "baremetal":
                custom_attributes_values = device.server.custom_attribute_data
            else:
                custom_attributes_values = device.custom_attribute_data
            if custom_attributes_values:
                custom_key = [str(i) for i in custom_attributes_values.keys()]
            for field in model_fields:
                if field == "name" and device.DEVICE_TYPE == "baremetal":
                    server_obj = device.server
                    row.append(validate_csv_value(unicode(server_obj.name) if server_obj else ""))
                    continue
                if device.DEVICE_TYPE == "baremetal":
                    device = device.server
                value = getattr(device, field, "")
                if field == "virtualization_type":
                    if hasattr(device, 'instance') and device.instance is not None:
                        value = getattr(device.instance, "virtualization_type", "")
                if field == "db_type" and device_type == "database_servers":
       
                    if hasattr(device, 'db_type'):
                        value = getattr(device.db_type, "name", "")
                
                if field == "model":
                    if hasattr(device, 'model'):
                        value = getattr(device.model, "name", "")

                
                if hasattr(value, "pk") and hasattr(value, "__unicode__"):
                    value = unicode(value)
                row.append(validate_csv_value(value))

            for key in custom_headers:
                row.append(
                    validate_csv_value((custom_attributes_values or {}).get(str(key), ''))
                )
                # row.append(validate_csv_value(attr.default_value or ""))
     
            # if device.DEVICE_TYPE == "baremetal":
            #     row.append(base64.urlsafe_b64encode(bm_device.uuid.bytes))
            # else:
            row.append(base64.urlsafe_b64encode(device.uuid.bytes))
            writer.writerow(row)

    return full_path


def update_device_fields(reader, headers, sheet_name, device_model_class, customer):
    """
    Update device fields from normalized_data, ignoring FKs and invalid attrs.
    """
    updated_devices = []
    failed_devices = []
    model_map = {
        "switch": (SwitchModel, Manufacturer),
        "firewall": (FirewallModel, Manufacturer),
        "load_balancer": (LoadBalancerModel, Manufacturer),
        "hypervisor": (ServerModel, ServerManufacturer),
        "baremetal": (ServerModel, ServerManufacturer),
        "mac_device": (ServerModel, ServerManufacturer),
        "storage":(StorageModel, StorageManufacturer)
    }
    model_fields = HEADERS_DICT[sheet_name]
    allowed_fields = HEADERS_DICT.get(sheet_name, [])
    for row in reader:
        try:
            normalized_data = {}
            if not row:
                return {}
            device_uuid = row[-1]
            if device_uuid:
                device_uuid = uuid.UUID(bytes=base64.urlsafe_b64decode(row[-1]))
            try:
                device = device_model_class.objects.get(uuid=device_uuid)
                if device.DEVICE_TYPE == "baremetal":
                    if getattr(device, 'server', None):
                        device_name = device.server.name
                elif device.DEVICE_TYPE == "database":
                    field = 'db_instance_name'
                    if getattr(device, 'db_instance_name', None):
                        device_name = device.db_instance_name
                else:
                    device_name = device.name
            except device_model_class.DoesNotExist:
                failed_devices.append(device_uuid)
                continue

            raw_row_map = dict(zip(headers, row))
            for idx, field in enumerate(model_fields):
                if idx < len(headers):
                    normalized_data[field] = row[idx] if row[idx] else ""
            ct = ContentType.objects.get_for_model(device_model_class)
            custom_attrs = list(CustomAttribute.objects.filter(
                        content_type=ct, customer=customer
                    ))
            asset_tag = normalized_data.pop('asset_tag', None)
            if hasattr(device, 'asset_tag'):
                device.asset_tag = asset_tag if asset_tag else device.asset_tag
                device.save()
            custom_attrs = CustomAttribute.objects.filter(customer=customer, content_type=ct)
            custom_json = {}
            for attr in custom_attrs:
                key = u"{}".format(attr.name)
                for h, v in raw_row_map.items():
                    if h.lower().startswith(attr.name.lower()):
                        normalized_data[key] = v
                        break
                else:
                    normalized_data[key] = ""
            for attr in custom_attrs:
                if str(attr.value_type) == 'Boolean':
                    val = normalized_data.get(str(attr.name), "").strip()
                    if val in ['true', 'TRUE']:
                        custom_json[str(attr.name)] = True
                    elif val in ['false', 'FALSE']:
                        custom_json[str(attr.name)] = False
                    else:
                        custom_json[str(attr.name)] = None
                if str(attr.value_type) == 'Integer':
                    try:
                        custom_json[str(attr.name)] = int(normalized_data[str(attr.name)])
                    except ValueError:
                        continue
                if str(attr.value_type) == 'Choice':
                    if normalized_data.get(str(attr.name), "") in [c for c in attr.choice_values]:
                        custom_json[str(attr.name)] = normalized_data[str(attr.name)]
                if str(attr.value_type) == 'Char':
                    custom_json[str(attr.name)] = normalized_data[str(attr.name)]
            if custom_json:
                if device.DEVICE_TYPE == 'baremetal':
                    server = device.server
                    server.custom_attribute_data = custom_json
                    server.save()
                else:
                    device.custom_attribute_data = custom_json
                    device.save()
            skip_fields = {
                'datacenter', 'cabinet', 'model',
                'collector', 'private_cloud', 'cloud',
                'manufacturer', 'asset_tag', 'os'
            }
            datacenter = normalized_data.pop('datacenter', None)
            cabinet = normalized_data.pop('cabinet', None)
            model = normalized_data.get('model', None)

            try:
                collector = normalized_data.pop('collector', None)
                os = normalized_data.pop('os', None)
                device.collector = AgentConfig.objects.get(name=collector)
            except AgentConfig.DoesNotExist:
                pass
            if datacenter:
                private_cloud = normalized_data.pop('cloud', [])
                datacenter = ColoCloud.objects.get(name=datacenter, customer=customer)
                device.datacenter = datacenter
                if cabinet:
                    cabinet = Cabinet.objects.get(name=cabinet)
                    for dc in cabinet.colocloud_set.all():
                        if dc.uuid == datacenter.uuid:
                            device.cabinet = cabinet
                            break
            if device.DEVICE_TYPE in ["baremetal", "hypervisor"]:
                private_cloud = normalized_data.pop('private_cloud', [])
            elif device.DEVICE_TYPE in ["storage","mac_device"]:
                private_cloud = normalized_data.pop('private_cloud',"")
            else:
                private_cloud = normalized_data.pop('cloud', [])

            if device.DEVICE_TYPE == "database":
                db_type = normalized_data.get("db_type","")
                db_type_obj = DatabaseType.objects.get(name = db_type)
                device.db_type = db_type_obj

            if private_cloud:
                if device.DEVICE_TYPE in ["storage","mac_device"]:
                    private_cloud_obj = PrivateCloud.objects.get(name= private_cloud)
                    device.private_cloud = private_cloud_obj


            if device.DEVICE_TYPE in model_map:
                manufacturer = normalized_data.get('manufacturer', None)
                model_class, manufacturer_class = model_map[device.DEVICE_TYPE]
                if manufacturer:
                    try:
                        manufacturer_obj = manufacturer_class.objects.get(name=manufacturer)
                        model_obj = model_class.objects.get(manufacturer=manufacturer_obj, name=model)
                        device.model = model_obj
                        device.manufacturer = manufacturer_obj
                    except manufacturer_class.DoesNotExist:           
                        manufacturer_obj = None
                        failed_devices.append(device_name)
                    except model_class.DoesNotExist:                 
                        model_obj = None
                        failed_devices.append(device_name)
            for field in allowed_fields:
                if field in skip_fields:
                    continue
                if field in normalized_data:
                    if field == "name":
                        device_name = normalized_data[field]
                        if device.DEVICE_TYPE == "baremetal":
                            server_obj = getattr(device, 'server', None)
                            if server_obj:
                                server_obj.name = normalized_data[field]
                                
                                try:
                                    server_obj.save()
                                except Exception:
                                    failed_devices.append(device_name)
                                    continue
                        pass
                    value = normalized_data[field]
                    try:
                        setattr(device, field, value)
                        device.save()
                        updated_devices.append(device_name)
                    except Exception:
                        continue
        except Exception as e:
            continue
    response_dict = {
        "updated_devices": set(updated_devices),
        "failed_devices": set(failed_devices)
    }
    return response_dict