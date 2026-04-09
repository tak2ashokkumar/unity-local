import uuid
import json
import csv
import tempfile
import os
import io

from copy import deepcopy
from tablib import Dataset
from data_book import ExcelDataBook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname, get_column_letter

from django.http import HttpResponse
from django.utils.dateparse import parse_datetime
from django.utils.http import urlquote
from rest_framework.authentication import SessionAuthentication, TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet
from rest_framework.decorators import detail_route, list_route

from rest_framework import generics
from rest_framework import status
from rest_framework.response import Response

from rest.customer.serializers import (
    CustomerSwitchCRUDSerializer, CustomerFirewallCRUDSerializer,
    CustomerLoadBalancerCRUDSerializer, CustomerHypervisorCRUDSerializer,
    CustomerBMSCRUDSerializer, CustomerStorageDeviceCRUDSerializer,
    CustomerMacDeviceCRUDSerializer, CustomerPDUCRUDSerializer,
    CustomerMobileDeviceCRUDSerializer, CustomerColoCloudSerializer,
    CustomerCabinetSerializer, CustomerDatabaseSerializer,
    CustomerSwitchPartialUpdateSerializer
)

from app.common.utils import OnboardStatus
from app.inventory.models import Firewall, Switch, LoadBalancer, BMServer, CustomAttribute
from app.customer_onboarding.utils import DEVICE_MODEL_MAP, HEADERS_DICT, DEVICE_MISMATCH_MAP
from integ.ObserviumBackend.tasks import activate_observium_monitoring
from cloud.vmware.models import VMwareVcenter
from integ.proxy.models import ReverseProxyCommon

from .resources import *
from .export_resources import (
    ExportColoCloudResource, ExportCabinetResource, VMWarePrivateCloudResource,
    ExportSwitchResource, ExportPDUResource, ExportFirewallResource, ExportLoadbalancerResource,
    ExportHyperVisorResource, ExportBMResource
)
from .models import OnboardingExcelFileData, NetworkScan
from .serializers import NetworkScanSerializer, OnboardingExcelFileDataSerializer
from .tasks import startScanning
from .utils import export_device_csv, update_device_fields

CloudResorceMap = {
    'VMware vCenter': VMWareVMResource,
    'VMware vCloud': VCloudVMResource,
    'OpenStack': OpenStackVMResource,
    'Proxmox': ProxmoxVMResource
}

COL_DATA_DICT_MAP = {
    'A': 'BM Type',
    'B': 'Cabinet Type',
    'C': 'Cabinet Renewal',
    'D': 'PDU Model',
    'E': 'Make',
    'F': 'Firewall Model',
    'G': 'Load balancer Model',
    'H': 'Switch Model',
    'I': 'Operating System',
    'J': 'OS Version',
    'K': 'Mac Mini OS',
    'L': 'Mac Mini Manufacturer',
    'M': 'Mac Mini Model',
    'N': 'Storage OS',
    'O': 'Storage Manufacturer',
    'P': 'Storage Model',
    'Q': 'Mobile Device Model',
    'R': 'Cloud',
    'S': 'Cabinet Model',
    'T': 'PDU Type',
    'U': 'Virtualization Type',
    'V': 'Device Type',
    'W': 'Power Circuit',
    'X': 'BMS Model',
    'Y': 'Hypervisor Model',
    'Z': 'Platform type',
    'AA': 'Datacenter',
    'AB': 'PDU Make',
    'AC': 'Database Type',
    'AD': 'Collector'
}


# Onboarding devices ViewSet
class OnboardingDevicesAuthentication:
    authentication_classes = (SessionAuthentication, TokenAuthentication)
    permission_classes = (IsAuthenticated,)

Model_Dict = {
    'Cabinets': Cabinet,
    'Datacenters': ColoCloud,
    'PDUs': PDU,
    'Firewalls': Firewall,
    'Load_Balancers': LoadBalancer,
    'Switches': Switch,
    'Hypervisors': Server,
    'Bare_Metals': BMServer,
    'Storage': StorageDevice,
    'MAC_Mini': MacDevice,
    'Mobile_Devices': MobileDevice,
    'Databases': DatabaseServer
}


class OnboardingDevicesBaseCreateView(OnboardingDevicesAuthentication, APIView):
    @property
    def model(self):
        return self.serializer_class.Meta.model

    def get(self, request, *args, **kwargs):
        return Response([], status=status.HTTP_200_OK)

    def post(self, request, *args, **kwargs):
        record_status = {}
        ctx = {'request': self.request}
        records = request.data
        all_valid = True
        try:
            onboardingexcel_objs = OnboardingExcelFileData.objects.filter(
                customer=self.request.user.org
            )
            for key, value in Model_Dict.iteritems():
                if value == self.model:
                    device_category = key
            if 'bmc_type' in records[0].keys():
                device_category = 'Bare_Metals'
            for record in records:
                unique_id = record['unique_id']
                if self.model == ColoCloud:
                    record['customer'] = self.request.user.organization_id
                elif self.model == Cabinet:
                    record['tags'] = []
                    if record['datacenter']:
                        record['colocloud_set'] = [
                            {'uuid': record['datacenter']}
                        ]
                if self.model not in [ColoCloud, Cabinet, MobileDevice, DatabaseServer]:
                    record['ip_address'] = record['snmp_ip']
                serializer = self.serializer_class(
                    instance=None,
                    data=record,
                    context=ctx
                )
                if serializer.is_valid():
                    if self.model == MobileDevice:
                        device_tagged = record["tagged_device"]
                        if device_tagged is not None:
                            model_name = 'server' if record["platform"] == 'Android' else 'macdevice'
                            content_type = ContentType.objects.get(model=model_name)
                            model_class = content_type.model_class()
                            device_object = model_class.objects.get(pk=device_tagged[0]['id'])
                            tagged_device = device_object
                            
                            device = serializer.save(tagged_device=tagged_device)
                        else:
                            device = serializer.save()
                    else:
                        device = serializer.save()
                    record_status[unique_id] = {
                        'saved': True,
                        'db_id': device.pk
                    }
                    for obj in onboardingexcel_objs:
                        # if isinstance(obj.onb_data, list):
                        if obj.onb_data and device_category in obj.onb_data:
                            for item in obj.onb_data[device_category]:
                                if item and 'unique_id' in item and item['unique_id'] == unique_id:
                                    obj.onb_data[device_category].remove(item)
                                    response_dict = prepare_json_data(
                                        device,
                                        device_category,
                                        unique_id
                                    )
                                    obj.onb_data[device_category].append(
                                        response_dict
                                    )
                            status_device_category = device_category if device_category in obj.onb_status else device_category.replace('_', ' ')
                            obj.onb_status[status_device_category]['success'] += 1
                            obj.save()
                else:
                    record_status[unique_id] = serializer.errors
                    all_valid = False

            # Updating json data obj for onboarding status and count
            for unique_id, value in record_status.iteritems():
                if 'saved' not in value:
                    for obj in onboardingexcel_objs:
                        # if isinstance(obj.onb_data, list):
                        if obj.onb_data and device_category in obj.onb_data:
                            for item in obj.onb_data[device_category]:
                                if item and 'unique_id' in item and item['unique_id'] == unique_id:
                                    item['onboarding_status'] = 'Failed'
                                    item['db_id'] = None
                            obj.save()

            if all_valid:
                return Response(record_status, status=status.HTTP_200_OK)
            else:
                return Response(
                    record_status,
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as error:
            logger.debug(error.message)
            return Response(
                {'error': error.message},
                status=status.HTTP_400_BAD_REQUEST
            )


class OnboardingSwitchesCreateView(OnboardingDevicesBaseCreateView):
    serializer_class = CustomerSwitchCRUDSerializer


class OnboardingFirewallsCreateView(OnboardingDevicesBaseCreateView):
    serializer_class = CustomerFirewallCRUDSerializer


class OnboardingLoadBalancersCreateView(OnboardingDevicesBaseCreateView):
    serializer_class = CustomerLoadBalancerCRUDSerializer


class OnboardingStorageDevicesCreateView(OnboardingDevicesBaseCreateView):
    serializer_class = CustomerStorageDeviceCRUDSerializer


class OnboardingServersCreateView(OnboardingDevicesBaseCreateView):
    serializer_class = CustomerHypervisorCRUDSerializer


class OnboardingPDUSCreateView(OnboardingDevicesBaseCreateView):
    serializer_class = CustomerPDUCRUDSerializer


class OnboardingBMSCreateView(OnboardingDevicesBaseCreateView):
    serializer_class = CustomerBMSCRUDSerializer


class OnboardingMacDeviceCreateView(OnboardingDevicesBaseCreateView):
    serializer_class = CustomerMacDeviceCRUDSerializer


class OnboardingMobileDeviceCreateView(OnboardingDevicesBaseCreateView):
    serializer_class = CustomerMobileDeviceCRUDSerializer


class OnboardingColoCloudCreateView(OnboardingDevicesBaseCreateView):
    serializer_class = CustomerColoCloudSerializer


class OnboardingCabinetCreateView(OnboardingDevicesBaseCreateView):
    serializer_class = CustomerCabinetSerializer


class OnboardingDatabaseCreateView(OnboardingDevicesBaseCreateView):
    serializer_class = CustomerDatabaseSerializer


class CustomerOnboardingView(APIView):

    def get(self, request, *args, **kwargs):
        file_path = 'media/onboarding_files/sample/OnboardingTemplate.xlsx'
        libre_office_file_copy = openpyxl.reader.excel.load_workbook(
            'app/customer_onboarding/OnboardingFileWithValidation.xlsx'
        )
        libre_office_file = openpyxl.reader.excel.load_workbook(
            'app/customer_onboarding/OnboardingFileWithValidation.xlsx'
        )

        # Dynamically adding make and model data into DATA sheet
        data_pdumodel = PDUModel.objects.all().values('model_number')
        data_make = Manufacturer.objects.all().values('name')
        data_fwmodel = FirewallModel.objects.all().values('name')
        data_lbmodel = LoadBalancerModel.objects.all().values('name')
        data_switchmodel = SwitchModel.objects.all().values('name')
        data_os = OperatingSystem.objects.all().values('name', 'version')
        data_macmini_os = OperatingSystem.objects.filter(platform_type='MacOS').values('name', 'version')
        data_mac_manufacturer = ServerManufacturer.objects.all().values('name')
        data_mac_model = ServerModel.objects.all().values('name')
        data_storage_os = OperatingSystem.objects.all().values('name', 'version')
        data_storage_manufacturer = StorageManufacturer.objects.all().values('name')
        data_storage_model = StorageModel.objects.all().values('name')
        data_mobile_device_model = MobileDeviceModel.objects.all().values('name')
        data_private_cloud = PrivateCloud.objects.filter(customer=request.user.org).values('name')
        data_power_circuit = PowerCircuit.objects.all()
        data_datacenters = ColoCloud.objects.filter(customer=request.user.org).values('name')
        data_pdu_manufacturer = PDUManufacturer.objects.all().values('name')
        data_collectors = AgentConfig.objects.filter(customer=request.user.org).values('name')
        # Fetching dynamic data length from DATA Tab
        for each_worksheet in libre_office_file.worksheets:
            copy_sheet = libre_office_file_copy[each_worksheet.title]
            if copy_sheet.title == "DATA":
                power_cicuit_list = []
                for data in data_power_circuit:
                    # power_cicuit_list.append(data.name + " (" + data.voltagetype.voltage_type + "/" + data.ampstype.amps_type + ")")
                    power_cicuit_list.append(data.name)
                # row_count = 3
                # for i in power_cicuit_list:
                #     copy_sheet.cell(row=row_count, column=23).value = i
                #     row_count += 1

                datasheet_dict = {
                    23: {'query': power_cicuit_list},
                    4: {'query': data_pdumodel, 'field_name': 'model_number'},
                    5: {'query': data_make, 'field_name': 'name'},
                    6: {'query': data_fwmodel, 'field_name': 'name'},
                    7: {'query': data_lbmodel, 'field_name': 'name'},
                    8: {'query': data_switchmodel, 'field_name': 'name'},
                    9: {'query': data_os, 'field_name': 'name'},
                    10: {'query': data_os, 'field_name': 'version'},
                    11: {'query': data_macmini_os},
                    12: {'query': data_mac_manufacturer, 'field_name': 'name'},
                    13: {'query': data_mac_model, 'field_name': 'name'},
                    14: {'query': data_storage_os},
                    15: {'query': data_storage_manufacturer, 'field_name': 'name'},
                    16: {'query': data_storage_model, 'field_name': 'name'},
                    17: {'query': data_mobile_device_model, 'field_name': 'name'},
                    18: {'query': data_private_cloud, 'field_name': 'name'},
                    24: {'query': data_mac_model, 'field_name': 'name'},
                    25: {'query': data_mac_model, 'field_name': 'name'},
                    28: {'query': data_pdu_manufacturer, 'field_name': 'name'},
                    27: {'query': data_datacenters, 'field_name': 'name'},
                    30: {'query': data_collectors, 'field_name': 'name'},
                }

                for column, details in datasheet_dict.items():
                    row_count = 3
                    query = details['query']
                    for item in query:
                        if column in (11, 14):
                            value = str(item['name']) + ' ' + str(item['version'])
                            copy_sheet.cell(row=row_count, column=column).value = value
                        elif column == 23:
                            copy_sheet.cell(row=row_count, column=column).value = item
                        else:
                            field_value = details['field_name']
                            copy_sheet.cell(row=row_count, column=column).value = item[field_value]
                        row_count += 1

                # Fetching column length from DATA sheet
                DATA_COL_Length_DICT = dict()
                # for col in range(1, copy_sheet.max_column + 1):
                for col in range(1, len(COL_DATA_DICT_MAP) + 1):
                    col_letter = get_column_letter(col)
                    max_col_row = len([cell for cell in copy_sheet[col_letter] if cell.value])
                    DATA_COL_Length_DICT[COL_DATA_DICT_MAP[col_letter]] = max_col_row
        # Data Validation dictionary to map dropdowns from DATA sheet to other device sheets
        validation_formula_dict = {
            "Firewalls": {
                "source_col": ["!${}$3:${}$".format(item, item) for item in ('E', 'R', 'F', 'AD', 'AA')],
                "source_header": ['Make', 'Cloud', 'Firewall Model', 'Collector', 'Datacenter'],
                "sheet_col": ["{}3:{}1048576".format(item, item) for item in ('C', 'F', 'D', 'L', 'E')]
            },
            "PDUs": {
                "source_col": ["!${}$3:${}$".format(item, item) for item in ('W', 'T', 'D', 'E', 'AD')],
                "source_header": ['Power Circuit', 'PDU Type', 'PDU Model', 'PDU Make', 'Collector'],
                "sheet_col": ["{}3:{}1048576".format(item, item) for item in ('D', 'E', 'H', 'G', 'M')]
            },
            "Cabinets": {
                "source_col": ["!${}$3:${}$".format(item, item) for item in ('C', 'S', 'AA')],
                "source_header": ['Cabinet Renewal', 'Cabinet Model', 'Datacenter'],
                "sheet_col": ["{}3:{}1048576".format(item, item) for item in ('I', 'D', 'C')]
            },
            "Switches": {
                "source_col": ["!${}$3:${}$".format(item, item) for item in ('E', 'R', 'H', 'AD', 'AA')],
                "source_header": ['Make', 'Cloud', 'Switch Model', 'Collector','Datacenter'],
                "sheet_col": ["{}3:{}1048576".format(item, item) for item in ('C', 'F', 'D', 'L', 'E')]
            },
            "Load Balancers": {
                "source_col": ["!${}$3:${}$".format(item, item) for item in ('E', 'R', 'G', 'AD', 'AA')],
                "source_header": ['Make', 'Cloud', 'Load balancer Model', 'Collector', 'Datacenter'],
                "sheet_col": ["{}3:{}1048576".format(item, item) for item in ('C', 'F', 'D', 'L', 'E')]
            },
            "Hypervisors": {
                "source_col": ["!${}$3:${}$".format(item, item) for item in ('Y', 'L', 'R', 'AD','AA', 'U', 'N')],
                "source_header": ['Hypervisor Model', 'Mac Mini Manufacturer', 'Cloud', 'Collector', 'Datacenter', 'Virtualization Type', 'Storage OS'],
                "sheet_col": ["{}3:{}1048576".format(item, item) for item in ('D', 'C', 'F', 'R', 'E', 'N', 'O')]
            },
            "Bare Metals": {
                "source_col": ["!${}$3:${}$".format(item, item) for item in ('X', 'L', 'R', 'AD', 'AA', 'N')],
                "source_header": ['BMS Model', 'Mac Mini Manufacturer', 'Cloud', 'Collector', 'Datacenter', 'Storage OS'],
                "sheet_col": ["{}3:{}1048576".format(item, item) for item in ('D', 'C', 'F', 'R', 'E', 'N')]
            },
            "Storage": {
                "source_col": ["!${}$3:${}$".format(item, item) for item in ('O', 'R', 'P', 'N', 'AD', 'AA')],
                "source_header": ['Storage Manufacturer', 'Cloud', 'Storage Model', 'Storage OS', 'Collector', 'Datacenter'],
                "sheet_col": ["{}3:{}1048576".format(item, item) for item in ('C', 'F', 'D', 'J', 'M', 'E')]
            },
            "MAC Mini": {
                "source_col": ["!${}$3:${}$".format(item, item) for item in ('L', 'R', 'M', 'K' ,'AD', 'AA')],
                "source_header": ['Storage Manufacturer', 'Cloud', 'Storage Model', 'Storage OS', 'Collector', 'Datacenter'],
                "sheet_col": ["{}3:{}1048576".format(item, item) for item in ('C', 'F', 'D', 'H', 'P', 'E')]
            },
            "Mobile Devices": {
                "source_col": ["!${0}$3:${1}$".format(item, item) for item in ('Z', 'AD', 'Q', 'AA')],
                "source_header": ['Platform type', 'Collector', 'Mobile Device Model', 'Datacenter'],
                "sheet_col": ["{}3:{}1048576".format(item, item) for item in ('F', 'I', 'C', 'D')]
            },
            "Databases": {
                "source_col": ["!${}$3:${}$".format(item, item) for item in ('AC',)],
                "source_header": ['Database Type'],
                "sheet_col": ["{}3:{}1048576".format(item, item) for item in ('B',)]
            },
        }
        

        for each_worksheet in libre_office_file.worksheets:
            copy_sheet = libre_office_file_copy[each_worksheet.title]
            data_validations = each_worksheet.data_validations.dataValidation
            for each_dv in data_validations:
                # 1048576 is used as this validation needs to applied to the whole column
                # This will result in a format C3:C1048576 if the coulmn label is C
                new_range = "{0}:{1}".format(str(each_dv.sqref)[:2], str(each_dv.sqref)[0] + '1048576')
                # removing the old range if present
                each_dv.sqref = ''
                copy_sheet.add_data_validation(each_dv)
                # Re adding the validation for whole row
                each_dv.add(new_range)
            # As we have only 1 line of dummy data in libre file
            if copy_sheet.title != "DATA":
                copy_sheet.delete_rows(3, 1)

            # Adding existing Datacenters in Datacenter sheet
            if copy_sheet.title == "Datacenters":
                row_count = 3
                for item in data_datacenters:
                    copy_sheet.cell(row=row_count, column=1).value = item['name']
                    row_count += 1

            # Adding data validation to each sheet columns from DATA sheet
            for device, details in validation_formula_dict.items():
                if copy_sheet.title == device:
                    for index in range(len(details["source_col"])):
                        field_name = details["source_header"][index]
                        if field_name == 'Datacenter':
                            formula = "Datacenters" + "!$A$3:$A$40"
                            data_val = DataValidation(
                                type="list", formula1=formula
                            )
                        else:
                            data_length = str(DATA_COL_Length_DICT[field_name])
                            if copy_sheet.title == "Databases":
                                data_length = str(DATA_COL_Length_DICT[field_name]+1)
                            formula = "DATA" + details["source_col"][index] + data_length
                            data_val = DataValidation(
                                type="list", formula1=formula
                            )
                        copy_sheet.add_data_validation(data_val)
                        data_val.add(details["sheet_col"][index])

        libre_office_file_copy.save(file_path)
        response = {
            'file_path': file_path
        }
        return Response(response, status=status.HTTP_200_OK)

    def post(self, request, *args, **kwargs):
        try:
            data_book = ExcelDataBook()
            uploaded_file = request.FILES.get('onboarding_file')
            customer_id = request.user.org.id
            customer = request.user.org
            onb_status = OnboardStatus()
            onboarding_status = True
            network_device_headers = [
                'name', 'management_ip', 'manufacturer',
                'model', 'datacenter', 'cloud',
                'cabinet', 'size', 'position',
                'asset_tag', 'snmp_ip', 'collector', 'onboarding_status'
            ]
            other_common_headers = [
                'name', 'management_ip', 'manufacturer',
                'model', 'datacenter', 'private_cloud',
                'cabinet', 'size', 'position'
            ]
            server_headers = ['num_cpus', 'num_cores', 'memory_mb', 'capacity_gb']
            headers_dict = {
                "Datacenters": [
                    'name', 'location', 'onboarding_status'
                ],
                "Cabinets": [
                    'name', 'size', 'datacenter',
                    'model', 'contract_start_date',
                    'contract_end_date', 'cost',
                    'annual_escalation', 'renewal',
                    'onboarding_status'
                ],
                "PDUs": [
                    'name', 'cabinet', 'size',
                    'power_circuit', 'pdu_type',
                    'position', 'manufacturer', 'model',
                    'ip_address', 'asset_tag', 'snmp_ip',
                    'sockets', 'collector', 'onboarding_status'
                ],
                "Firewalls": network_device_headers,
                "Switches": network_device_headers,
                "Load Balancers": network_device_headers,
                "Hypervisors": other_common_headers + server_headers + [
                    'virtualization_type', 'os', 'asset_tag',
                    'snmp_ip',  'collector', 'onboarding_status'
                ],
                "Bare Metals": other_common_headers + server_headers + [
                    'os', 'bmc_type', 'asset_tag',
                    'snmp_ip',  'collector', 'onboarding_status'
                ],
                "Storage": other_common_headers + [
                    'os', 'asset_tag', 'snmp_ip',  'collector','onboarding_status'
                ],
                "Mobile Devices": [
                    'name', 'serial_number', 'model',
                    'datacenter', 'ip_address', 'platform',
                    'device_type', 'tagged_device', 'collector', 'onboarding_status'
                ],
                "MAC Mini": [
                    'name', 'serial_number', 'manufacturer',
                    'model', 'datacenter', 'private_cloud',
                    'cabinet', 'os', 'management_ip',
                    'num_cpus', 'num_cores', 'memory_mb',
                    'capacity_gb', 'asset_tag', 'snmp_ip',
                    'collector', 'onboarding_status'
                ],
                "Databases": ['name', 'db_type', 'port', 'management_ip']
            }

            if uploaded_file:
                open_data_sheet = openpyxl.reader.excel.load_workbook(uploaded_file)
                required_sheets = headers_dict.keys()
                missing_sheets = [sheet for sheet in required_sheets if sheet not in open_data_sheet.sheetnames]
                if missing_sheets:
                    raise Exception("Upload a Valid file")
                onboarding_file = OnboardingExcelFileData.objects.create(
                    customer_id=customer_id,
                    user=request.user,
                    document=uploaded_file
                )
                file_object = onboarding_file.document
                file_url = file_object.path
                json_sheets = {}
                count_devices = {}

                data_sheet = open_data_sheet.get_sheet_by_name("DATA")
                existing_dc_count = 0
                for row in range(3, data_sheet.max_row):
                    if(data_sheet.cell(row, 30).value is None):
                        break
                    existing_dc_count += 1

                for each_sheet in open_data_sheet.worksheets:
                    for device, headers in headers_dict.items():
                        if each_sheet.title == device:
                            device_list = []
                            flag_row = True
                            row_count = 1
                            while flag_row:
                                if each_sheet.cell(row=row_count, column=1).value is None:
                                    break
                                else:
                                    row_count += 1
                            for row in range(3, row_count):
                                # Skipping existing datacenter values added from DATA sheet
                                if device == "Datacenters":
                                    if row in range(3, existing_dc_count + 3):
                                        continue
                                json_dict = {}
                                for i in range(1, len(headers) + 1):
                                    if headers[i - 1] != 'onboarding_status':
                                        data_value = each_sheet.cell(row=row, column=i).value
                                        if each_sheet.title == "Cabinets":
                                            if (
                                                headers[i - 1] == 'contract_start_date' or
                                                headers[i - 1] == 'contract_end_date'
                                            ):
                                                json_dict[str(headers[i - 1])] = str(data_value)\
                                                    if data_value else None
                                            else:
                                                json_dict[str(headers[i - 1])] = data_value
                                        else:
                                            json_dict[str(headers[i - 1])] = data_value
                                    else:
                                        json_dict[str(headers[i - 1])] = None

                                json_dict['unique_id'] = str(generate_uuid())
                                device_list.append(json_dict)
                            if each_sheet.title in ("Load Balancers", "Bare Metals", "Mobile Devices", "MAC Mini"):
                                sheet_key = each_sheet.title.replace(" ", "_")
                            else:
                                sheet_key = each_sheet.title
                            json_sheets[sheet_key] = device_list
                            count_devices[each_sheet.title] = {
                                'count': len(device_list), 'success': 0, 'failed': 0
                            }

                onboarding_file.onb_data = json_sheets
                onboarding_file.onb_status = count_devices
                onboarding_file.save()
            response = {
                'status': True
            }
            return Response(response, status=status.HTTP_200_OK)
        except Exception as error:
            response = {
                'status': False
            }
            logger.error(error)
            return Response(response, status=status.HTTP_400_BAD_REQUEST)


class EnableVMMonitoringView(APIView):
    def get(self, request, *args, **kwargs):
        cloud_type = self.request.query_params.get('cloud_type', None)
        if cloud_type:
            vm_resource = CloudResorceMap.get(cloud_type)()
            if cloud_type == 'Proxmox':
                query_filter = {'cluster__private_cloud__customer': request.user.org}
            else:
                query_filter = {'cloud__customer': request.user.org}
            query_set = vm_resource._meta.model.objects.filter(**query_filter).order_by('id')
            dataset = vm_resource.export(query_set)
            response = HttpResponse(
                dataset.xls,
                content_type='application/vnd.ms-excel'
            )
            file_name = "{}_vms.xls".format(cloud_type.replace(" ", "_"))
            response['Content-Disposition'] = 'attachment; filename={}'.format(file_name)
            return response
        else:
            return None

    def post(self, request, *args, **kwargs):
        cloud_type = self.request.query_params.get('cloud_type', None)
        if cloud_type:
            vm_resource = CloudResorceMap.get(cloud_type)()
            dataset = Dataset()
            vm_input_file = request.FILES['vm_file']

            imported_data = dataset.load(vm_input_file.read())
            result = vm_resource.import_data(dataset, dry_run=True)

            if not result.has_errors():
                vm_resource.import_data(dataset, dry_run=False)
                response = {
                    'success': True
                }
                customer = request.user.org
                task = activate_observium_monitoring.delay(customer.id)
                onb_status = OnboardStatus()
                new_status = {
                    "monitoring_start": True,
                    "monitoring_error": False
                }
                onb_status.update_db_status(customer, new_status)
                return Response(response, status=status.HTTP_200_OK)
            response = {
                'success': False
            }
            return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {
                'success': False
            }
            return Response(response, status=status.HTTP_400_BAD_REQUEST)


class NetworkScanViewSet(ModelViewSet):
    serializer_class = NetworkScanSerializer
    lookup_field = 'uuid'

    def get_queryset(self):
        queryset = NetworkScan.objects.filter(
            org=self.request.user.org
        ).order_by('-updated_at')

        return queryset

    def get(self, request, *args, **kwargs):
        return Response(NetworkScanSerializer(
            self.get_queryset(),
            many=True,
            context={'request': request}).data,
            status=status.HTTP_200_OK
        )

    def create(self, request, *args, **kwargs):
        try:
            inet = request.data.get('inet', None)
            new_scan_obj = NetworkScan.objects.create(
                org=request.user.org,
                inet=inet,
                scan_status=NetworkScan.SCAN_STATUS[0][0]
            )
            task = startScanning.delay(inet, new_scan_obj.id)
            return Response(NetworkScanSerializer(
                self.get_queryset(),
                many=True,
                context={'request': request}).data,
                status=status.HTTP_200_OK)
        except Exception as e:
            logger.error("error %s" % e)
            return Response('Something went wrong, please try again later', status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        try:
            scan = self.get_object()
            scan.scan_status = NetworkScan.SCAN_STATUS[2][0]
            scan.save()

            task = startScanning.delay(scan.inet, scan.id)
            return Response(NetworkScanSerializer(
                self.get_queryset(),
                many=True,
                context={'request': request}).data,
                status=status.HTTP_200_OK)
        except Exception as e:
            logger.error("error %s" % e)
            return Response('Something went wrong, please try again later', status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, uuid=None, *args, **kwargs):
        try:
            scan = self.get_object()
            scan.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Exception as error:
            logger.debug("error while deleting:%s ", error)
            return Response(status=status.HTTP_400_BAD_REQUEST)


class DownloadNetworkScans(generics.ListAPIView):

    def list(self, request, *args, **kwargs):
        try:
            uuid = request.GET.get('uuid', None)
            scan = NetworkScan.objects.get(uuid=uuid)
            results = json.loads(scan.scan_results)

            scan_done_at = scan.updated_at.strftime("%Y-%m-%d %H:%M:%S")
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="scan_report_%s.csv"' % (scan_done_at)
            writer = csv.writer(response)
            writer.writerow(['IP Address', 'Hostname', 'Status', 'Vendor', 'Operating System', 'Device Type'])
            for row in results:
                writer.writerow([
                    row['ip_address'],
                    row['hostname'] or 'N/A',
                    row['status'],
                    row['vendor'],
                    row['os'],
                    row['device_type'],
                ])
            return response
        except Exception as e:
            logger.error("error in downloadReport - %s" % e)
            return Response('Something went wrong, please try again later', status=status.HTTP_400_BAD_REQUEST)


class CustomerAssestDownloadView(APIView):
    def get(self, request, *args, **kwargs):
        libre_office_file = openpyxl.reader.excel.load_workbook(
            'app/customer_onboarding/OnboardingFileWithValidation.xlsx'
        )
        libre_office_file_copy = deepcopy(libre_office_file)

        for each_worksheet in libre_office_file.worksheets:
            copy_sheet = libre_office_file_copy[each_worksheet.title]
            data_validations = each_worksheet.data_validations.dataValidation
            for each_dv in data_validations:
                # 1048576 is used as this validation needs to applied to the whole column
                # This will result in a format C3:C1048576 if the coulmn label is C
                new_range = "{0}:{1}".format(str(each_dv.sqref)[:2], str(each_dv.sqref)[0] + '1048576')
                # removing the old range if present
                each_dv.sqref = ''
                copy_sheet.add_data_validation(each_dv)
                # Re adding the validation for whole row
                each_dv.add(new_range)
            # As we have only 1 line of dummy data in libre file
            if copy_sheet.title != "DATA":
                copy_sheet.delete_rows(3, 1)

                for _cell in copy_sheet[1:1]:
                    _cell.alignment = Alignment(wrap_text=True)

            if copy_sheet.title == 'Datacenters':
                colo_clouds = ColoCloud.objects.filter(
                    customer=request.user.org
                )
                colo_exp_resource = ExportColoCloudResource()
                colo_cloud_dataset = colo_exp_resource.export(colo_clouds)
                for entry in colo_cloud_dataset._data:
                    copy_sheet.append(list(entry))

            elif copy_sheet.title == 'Cabinets':
                cabinets = Cabinet.objects.filter(
                    customers=request.user.org
                )
                cabinet_exp_resource = ExportCabinetResource()
                cabinet_dataset = cabinet_exp_resource.export(cabinets)
                for entry in cabinet_dataset._data:
                    copy_sheet.append(list(entry))

            # elif copy_sheet.title == 'vCenters':
            #     vcenters = VMwareVcenter.objects.filter(
            #         private_cloud__customer=request.user.org
            #     )
            #     vcenter_exp_resource = VMWarePrivateCloudResource()
            #     vcenter_dataset = vcenter_exp_resource.export(vcenters)
            #     for entry in vcenter_dataset._data:
            #         print "Inside entry", entry
            #         copy_sheet.append(list(entry))

            elif copy_sheet.title == 'PDUs':
                pdus = PDU.objects.filter(
                    cabinet__customers=request.user.org
                )
                pdu_exp_resource = ExportPDUResource()
                pdu_dataset = pdu_exp_resource.export(pdus)
                for entry in pdu_dataset._data:
                    copy_sheet.append(list(entry))

            elif copy_sheet.title == 'Firewalls':
                firewalls = Firewall.objects.filter(
                    customers=request.user.org,
                    is_shared=False
                )
                firewall_exp_resource = ExportFirewallResource()
                firewall_dataset = firewall_exp_resource.export(firewalls)
                for entry in firewall_dataset._data:
                    copy_sheet.append(list(entry))

            elif copy_sheet.title == 'Switches':
                switches = Switch.objects.filter(
                    customers=request.user.org,
                    is_shared=False
                )
                switch_exp_resource = ExportSwitchResource()
                switch_dataset = switch_exp_resource.export(switches)
                for entry in switch_dataset._data:
                    copy_sheet.append(list(entry))

            elif copy_sheet.title == 'Load Balancers':
                loadbalancers = LoadBalancer.objects.filter(
                    customers=request.user.org,
                    is_shared=False
                )
                lb_exp_resource = ExportLoadbalancerResource()
                lb_dataset = lb_exp_resource.export(loadbalancers)
                for entry in lb_dataset._data:
                    copy_sheet.append(list(entry))

            elif copy_sheet.title == 'Hypervisors':
                hypervisors = Instance.objects.filter(
                    system__customer=request.user.org
                )
                hypervisor_exp_resource = ExportHyperVisorResource()
                hv_dataset = hypervisor_exp_resource.export(hypervisors)
                for entry in hv_dataset._data:
                    copy_sheet.append(list(entry))

            elif copy_sheet.title == 'Bare Metals':
                bm_servers = BMServer.objects.filter(
                    server__customer=request.user.org
                )
                bm_server_exp_resource = ExportBMResource()
                bms_dataset = bm_server_exp_resource.export(bm_servers)
                for entry in bms_dataset._data:
                    copy_sheet.append(list(entry))

        response = HttpResponse(
            content=save_virtual_workbook(libre_office_file_copy),
            content_type='application/ms-excel'
        )
        file_name = "assets_download.xlsx"
        response['Content-Disposition'] = 'attachment; filename={}'.format(file_name)

        return response


class OnboardingExcelViewSet(ModelViewSet):
    serializer_class = OnboardingExcelFileDataSerializer
    lookup_field = 'uuid'

    def get_queryset(self):
        excel_list = self.request.query_params.getlist('uuid')
        if excel_list:
            queryset = OnboardingExcelFileData.objects.filter(
                customer=self.request.user.org, uuid__in=excel_list).order_by('-updated_at')
        else:
            queryset = OnboardingExcelFileData.objects.filter(
                customer=self.request.user.org).order_by('-updated_at')
        return queryset

    @detail_route(methods=['GET'])
    def get_result(self, request, *args, **kwargs):
        excel_result = self.get_object()
        device_type = request.query_params.get('type')
        response_data = validate_onbaording_data(excel_result.onb_data, excel_result.customer)
        excel_result.onb_data = response_data
        excel_result.save()
        if device_type in excel_result.onb_data:
            response = excel_result.onb_data[device_type]
        else:
            response = excel_result.onb_data
        return Response(response, status=status.HTTP_200_OK)

    @list_route(methods=['GET'])
    def get_all_result(self, request, *args, **kwargs):
        device_type = request.query_params.get('type')
        excel_list = self.request.query_params.getlist('uuid')
        excel_org_result = OnboardingExcelFileData.objects.filter(
            customer=request.user.org, uuid__in=excel_list)
        for excel_result in excel_org_result:
            response_data = validate_onbaording_data(excel_result.onb_data, request.user.org)
            excel_result.onb_data = response_data
            excel_result.save()
        if device_type:
            response_list = list()
            for excel_result in excel_org_result:
                if device_type in excel_result.onb_data:
                    response_list.extend(excel_result.onb_data[device_type])
            if device_type == 'Mobile_Devices':
                response_list = {v['serial_number']: v for v in response_list}.values()
            else:
                response_list = {v['name']: v for v in response_list if v and 'name' in v}.values()
            return Response(response_list, status=status.HTTP_200_OK)
        else:
            response_dict = dict()
            for excel_result in excel_org_result:
                for key in excel_result.onb_data.keys():
                    if key in response_dict.keys():
                        response_dict[key].extend(excel_result.onb_data[key])
                    else:
                        response_dict[key] = excel_result.onb_data[key]
            sorted_data_dict = dict()
            for key in response_dict.keys():
                sorted_data_dict[key] = {v['name']: v for v in response_dict[key]}.values()

            return Response(sorted_data_dict, status=status.HTTP_200_OK)

    @list_route(methods=['GET'])
    def total_summary(self, request, *args, **kwargs):
        try:
            # device_type = request.query_params.get('type')
            excel_org_result = OnboardingExcelFileData.objects.filter(
                customer=request.user.org)
            summary_dict = {"count": 0, "failed": 0, "success": 0}
            response_dict = {
                "Bare_Metals": summary_dict,
                "Cabinets": summary_dict,
                "Storage": summary_dict,
                "PDUs": summary_dict,
                "Switches": summary_dict,
                "Hypervisors": summary_dict,
                "MAC_Mini": summary_dict,
                "Mobile_Devices": summary_dict,
                "Datacenters": summary_dict,
                "Firewalls": summary_dict,
                "Load_Balancers": summary_dict,
                "Databases": summary_dict
            }
            for excel_result in excel_org_result:
                if excel_result.onb_status:
                    for key in excel_result.onb_status.keys():
                        if key in response_dict.keys():
                            count_dict = {'count': response_dict[key]['count'] + excel_result.onb_status[key]['count'],
                                          'success': response_dict[key]['success'] + excel_result.onb_status[key]['success'],
                                          'failed': response_dict[key]['failed'] + excel_result.onb_status[key]['failed']}
                            response_dict[key] = count_dict
                        else:
                            response_dict[key] = {'count': excel_result.onb_status[key]['count'],
                                                  'success': excel_result.onb_status[key]['success'],
                                                  'failed': excel_result.onb_status[key]['failed']}
            return Response(response_dict, status=status.HTTP_200_OK)
        except Exception as error:
            logger.error("Error while fetching summary:{}".format(error))
            return Response({'error': error.message}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, uuid=None, *args, **kwargs):
        try:
            excel = self.get_object()
            excel.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Exception as error:
            logger.debug("error while deleting:%s ", error)
            return Response(status=status.HTTP_400_BAD_REQUEST)

    @list_route(methods=['DELETE'])
    def delete_device(self, request):
        uuid = self.request.GET.get('uuid')
        device_type = self.request.GET.get('device_type')
        unique_id = self.request.GET.get('unique_id')
        try:
            onb_obj = OnboardingExcelFileData.objects.get(uuid=uuid)
            onb_data = onb_obj.onb_data
            if device_type in onb_data.keys():
                for data in onb_data[device_type]:
                    if data['unique_id'] == unique_id:
                        onb_data[device_type].remove(data)
                        onb_obj.onb_status[device_type.replace("_", " ")]['count'] -= 1
                        break
                onb_obj.save()
            return Response(status=status.HTTP_200_OK)
        except OnboardingExcelFileData.DoesNotExist:
            logger.error("Failed to delete the device")
            return Response(status=status.HTTP_400_BAD_REQUEST)

    @detail_route(methods=['PUT'])
    def delete_excel(self, request, *args, **kwargs):
        try:
            onb_obj = self.get_object()
        except OnboardingExcelFileData.DoesNotExist:
            logger.error("Failed to delete the device")
            return Response(status=status.HTTP_400_BAD_REQUEST)
        else:
            onb_obj.delete()
            return Response(status=status.HTTP_200_OK)

    @list_route(methods=['POST'])
    def onboard_next(self, request):
        try:
            device_list = request.data['device']
            device_type = request.data['device_type']
            onb_obj = OnboardingExcelFileData.objects.filter(customer=request.user.org)
            for onb_item in onb_obj:
                for device in device_list:
                    if hasattr(onb_item.onb_data, device_type):
                        for item in onb_item.onb_data[device_type]:
                            if item and 'unique_id' in item and device['unique_id'] == item['unique_id']:
                                updated_device, error = prepare_json_temp(device, device_type, device['unique_id'])
                                onb_item.onb_data[device_type].remove(item)
                                onb_item.onb_data[device_type].append(updated_device)
                                onb_item.save()
                                if error:
                                    raise Exception(error)
                                break
            return Response(status=status.HTTP_200_OK)
        except Exception as error:
            return Response({'error ': str(error)}, status=status.HTTP_400_BAD_REQUEST)

    @list_route(methods=['GET'])
    def get_bulk_upload_file(self, request):
        customer = request.user.org
        custom_attributes = CustomAttribute.objects.all()
        resource_types = self.request.query_params.get('device_type', [])
        device_uuids = self.request.query_params.getlist('uuids', 'all')
        custom_attributes = custom_attributes.filter(customer=self.request.user.org)
        file_path = export_device_csv(resource_types, device_uuids, customer)
        return Response({"file_path": file_path}, status=status.HTTP_200_OK)

    @list_route(methods=['GET'])
    def download_bulk_upload_file(self, request):
        file_path = request.GET.get('file_path')
        if not file_path or not os.path.exists(file_path):
            return HttpResponse("File not found", status=404)
        try:
            with open(file_path, 'rb') as f:
                data = f.read()
            os.remove(file_path)
            filename = os.path.basename(file_path)
            safe_filename = urlquote(filename)

            response = HttpResponse(
                data
            )
            response['Content-Disposition'] = 'attachment; filename="{}"'.format(safe_filename)
            return response

        except Exception as e:
            if os.path.exists(file_path):
                os.remove(file_path)
            return HttpResponse(str(e), status=500)

    @list_route(methods=['POST'])
    def upload_bulk_update(self, request, *args, **kwargs):
        customer = self.request.user.org
        device_type = request.data.get("device_type").lower()
        for k, v in DEVICE_MISMATCH_MAP.items():
            if device_type == v:
                device_type = k
                break

        if device_type not in DEVICE_MODEL_MAP.keys():
            return Response(
                {"status": False, "error": "Unsupported device type: {}".format(device_type)},
                status=status.HTTP_400_BAD_REQUEST
            )

        device_model_class, sheet_name = DEVICE_MODEL_MAP[device_type]
        model_fields = HEADERS_DICT[sheet_name]

        uploaded_file = request.FILES.get("device_bulkupdate_file")
        if not uploaded_file:
            return Response(
                {"status": False, "error": "CSV file is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        from io import TextIOWrapper
        csv_file = TextIOWrapper(uploaded_file.file, encoding="utf-8")
        reader = csv.reader(csv_file)
        headers = next(reader)
        response_dict = update_device_fields(reader, headers, sheet_name, device_model_class, customer)
        return Response(response_dict)

